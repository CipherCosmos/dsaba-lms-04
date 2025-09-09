from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, Reference
import io
from datetime import datetime
from typing import Dict, List, Any

def generate_pdf_report(report_data: Dict[str, Any]) -> bytes:
    """Generate PDF report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=12,
        textColor=colors.darkblue
    )
    
    elements = []
    
    # Title
    title = f"{report_data['report_type'].replace('_', ' ').title()} Report"
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))
    
    # Generated date
    date_text = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 24))
    
    if report_data['report_type'] == 'student_performance':
        generate_student_performance_pdf(elements, report_data, styles, heading_style)
    elif report_data['report_type'] == 'co_po_attainment':
        generate_co_po_attainment_pdf(elements, report_data, styles, heading_style)
    elif report_data['report_type'] == 'nba_compliance':
        generate_nba_compliance_pdf(elements, report_data, styles, heading_style)
    
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()

def generate_student_performance_pdf(elements, report_data, styles, heading_style):
    """Generate student performance section in PDF"""
    elements.append(Paragraph("Student Performance Analysis", heading_style))
    elements.append(Spacer(1, 12))
    
    # Summary statistics
    total_students = len(report_data['report_data'])
    avg_percentage = sum(s['percentage'] for s in report_data['report_data']) / total_students if total_students > 0 else 0
    pass_rate = len([s for s in report_data['report_data'] if s['percentage'] >= 50]) / total_students * 100 if total_students > 0 else 0
    
    summary_text = f"""
    <b>Summary Statistics:</b><br/>
    Total Students: {total_students}<br/>
    Average Percentage: {avg_percentage:.2f}%<br/>
    Pass Rate: {pass_rate:.2f}%<br/>
    """
    elements.append(Paragraph(summary_text, styles['Normal']))
    elements.append(Spacer(1, 18))
    
    # Student performance table
    data = [['Rank', 'Student Name', 'Student ID', 'Percentage', 'Grade']]
    
    # Sort by percentage for ranking
    sorted_students = sorted(report_data['report_data'], key=lambda x: x['percentage'], reverse=True)
    
    for i, student in enumerate(sorted_students[:20]):  # Top 20 students
        grade = get_grade_from_percentage(student['percentage'])
        data.append([
            str(i + 1),
            student['student_name'],
            student['student_id'],
            f"{student['percentage']:.1f}%",
            grade
        ])
    
    table = Table(data, colWidths=[0.5*inch, 2*inch, 1.5*inch, 1*inch, 0.8*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)
    elements.append(Spacer(1, 24))

def generate_co_po_attainment_pdf(elements, report_data, styles, heading_style):
    """Generate CO/PO attainment section in PDF"""
    elements.append(Paragraph("Course and Program Outcomes Attainment", heading_style))
    elements.append(Spacer(1, 12))
    
    for subject_data in report_data['report_data']:
        # Subject header
        subject_title = f"{subject_data['subject_name']} ({subject_data['subject_code']})"
        elements.append(Paragraph(subject_title, styles['Heading3']))
        elements.append(Spacer(1, 6))
        
        # CO Attainment table
        if subject_data['co_attainment']:
            elements.append(Paragraph("Course Outcomes (CO) Attainment:", styles['Normal']))
            co_data = [['Course Outcome', 'Attainment (%)', 'Status']]
            
            for co, attainment in subject_data['co_attainment'].items():
                status = "Achieved" if attainment >= 60 else "Not Achieved"
                co_data.append([co, f"{attainment:.2f}%", status])
            
            co_table = Table(co_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
            co_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(co_table)
            elements.append(Spacer(1, 12))
        
        # PO Attainment table
        if subject_data['po_attainment']:
            elements.append(Paragraph("Program Outcomes (PO) Attainment:", styles['Normal']))
            po_data = [['Program Outcome', 'Attainment (%)', 'Status']]
            
            for po, attainment in subject_data['po_attainment'].items():
                status = "Achieved" if attainment >= 60 else "Not Achieved"
                po_data.append([po, f"{attainment:.2f}%", status])
            
            po_table = Table(po_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
            po_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            elements.append(po_table)
            elements.append(Spacer(1, 18))

def generate_nba_compliance_pdf(elements, report_data, styles, heading_style):
    """Generate NBA compliance section in PDF"""
    elements.append(Paragraph("NBA Compliance Report", heading_style))
    elements.append(Spacer(1, 12))
    
    # Overall compliance status
    compliance_text = f"<b>Overall NBA Compliance Status:</b> {'COMPLIANT' if report_data['overall_compliance'] else 'NON-COMPLIANT'}"
    elements.append(Paragraph(compliance_text, styles['Normal']))
    elements.append(Spacer(1, 12))
    
    # Subject-wise compliance
    data = [['Subject', 'Subject Code', 'Compliance Status', 'Action Required']]
    
    for subject_data in report_data['report_data']:
        status = "✓ Compliant" if subject_data['compliance_status'] else "✗ Non-Compliant"
        action = "Yes" if subject_data['action_required'] else "No"
        
        data.append([
            subject_data['subject_name'],
            subject_data['subject_code'],
            status,
            action
        ])
    
    table = Table(data, colWidths=[2.5*inch, 1.2*inch, 1.5*inch, 1*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.red if not report_data['overall_compliance'] else colors.darkgreen),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(table)

def generate_excel_report(report_data: Dict[str, Any]) -> bytes:
    """Generate Excel report"""
    buffer = io.BytesIO()
    workbook = openpyxl.Workbook()
    
    if report_data['report_type'] == 'student_performance':
        generate_student_performance_excel(workbook, report_data)
    elif report_data['report_type'] == 'co_po_attainment':
        generate_co_po_attainment_excel(workbook, report_data)
    elif report_data['report_type'] == 'nba_compliance':
        generate_nba_compliance_excel(workbook, report_data)
    
    workbook.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

def generate_student_performance_excel(workbook, report_data):
    """Generate student performance Excel sheet"""
    ws = workbook.active
    ws.title = "Student Performance"
    
    # Headers
    headers = ['Rank', 'Student Name', 'Student ID', 'Percentage', 'Grade']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    sorted_students = sorted(report_data['report_data'], key=lambda x: x['percentage'], reverse=True)
    
    for row, student in enumerate(sorted_students, 2):
        ws.cell(row=row, column=1, value=row-1)  # Rank
        ws.cell(row=row, column=2, value=student['student_name'])
        ws.cell(row=row, column=3, value=student['student_id'])
        ws.cell(row=row, column=4, value=round(student['percentage'], 2))
        ws.cell(row=row, column=5, value=get_grade_from_percentage(student['percentage']))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Add summary sheet
    summary_ws = workbook.create_sheet("Summary")
    total_students = len(report_data['report_data'])
    avg_percentage = sum(s['percentage'] for s in report_data['report_data']) / total_students if total_students > 0 else 0
    pass_rate = len([s for s in report_data['report_data'] if s['percentage'] >= 50]) / total_students * 100 if total_students > 0 else 0
    
    summary_data = [
        ['Metric', 'Value'],
        ['Total Students', total_students],
        ['Average Percentage', f"{avg_percentage:.2f}%"],
        ['Pass Rate', f"{pass_rate:.2f}%"],
        ['Top Performer', sorted_students[0]['student_name'] if sorted_students else 'N/A'],
        ['Highest Score', f"{sorted_students[0]['percentage']:.2f}%" if sorted_students else 'N/A']
    ]
    
    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = summary_ws.cell(row=row_idx, column=col_idx, value=value)
            if row_idx == 1:  # Header row
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

def generate_co_po_attainment_excel(workbook, report_data):
    """Generate CO/PO attainment Excel sheets"""
    ws = workbook.active
    ws.title = "CO-PO Attainment Summary"
    
    # Headers
    headers = ['Subject', 'Subject Code'] + list(set([co for subject in report_data['report_data'] for co in subject.get('co_attainment', {}).keys()]))
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    
    # Data
    for row, subject_data in enumerate(report_data['report_data'], 2):
        ws.cell(row=row, column=1, value=subject_data['subject_name'])
        ws.cell(row=row, column=2, value=subject_data['subject_code'])
        
        for col_idx, co in enumerate(headers[2:], 3):
            attainment = subject_data.get('co_attainment', {}).get(co, 0)
            cell = ws.cell(row=row, column=col_idx, value=round(attainment, 2))
            
            # Color coding based on attainment
            if attainment >= 75:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
            elif attainment >= 60:
                cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

def generate_nba_compliance_excel(workbook, report_data):
    """Generate NBA compliance Excel sheet"""
    ws = workbook.active
    ws.title = "NBA Compliance"
    
    # Overall compliance status
    ws.cell(row=1, column=1, value="NBA COMPLIANCE REPORT")
    ws.cell(row=1, column=1).font = Font(size=16, bold=True)
    
    compliance_status = "COMPLIANT" if report_data['overall_compliance'] else "NON-COMPLIANT"
    ws.cell(row=2, column=1, value=f"Overall Status: {compliance_status}")
    ws.cell(row=2, column=1).font = Font(size=12, bold=True)
    
    if report_data['overall_compliance']:
        ws.cell(row=2, column=1).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    else:
        ws.cell(row=2, column=1).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    
    # Headers
    headers = ['Subject', 'Subject Code', 'Compliance Status', 'Action Required', 'CO Attainment Details']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Data
    for row, subject_data in enumerate(report_data['report_data'], 5):
        ws.cell(row=row, column=1, value=subject_data['subject_name'])
        ws.cell(row=row, column=2, value=subject_data['subject_code'])
        
        status_cell = ws.cell(row=row, column=3, value="Compliant" if subject_data['compliance_status'] else "Non-Compliant")
        if subject_data['compliance_status']:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        ws.cell(row=row, column=4, value="Yes" if subject_data['action_required'] else "No")
        
        # CO attainment details
        co_details = ", ".join([f"{co}: {att:.1f}%" for co, att in subject_data.get('co_attainment', {}).items()])
        ws.cell(row=row, column=5, value=co_details)

def get_grade_from_percentage(percentage):
    """Convert percentage to grade"""
    if percentage >= 90:
        return 'A+'
    elif percentage >= 80:
        return 'A'
    elif percentage >= 70:
        return 'B+'
    elif percentage >= 60:
        return 'B'
    elif percentage >= 50:
        return 'C'
    else:
        return 'F'