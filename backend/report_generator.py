import os
import io
import json
from datetime import datetime
from typing import Dict, List, Any, Optional
from sqlalchemy.orm import Session
from sqlalchemy import and_, func, desc
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from reportlab.lib.pagesizes import letter, A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart, PieChart
# from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties

from models import User, Subject, Exam, Question, Mark, Class, CODefinition, PODefinition, COPOMatrix, QuestionCOWeight
from crud import *
from database import SessionLocal

class ReportGenerator:
    def __init__(self, db: Session):
        self.db = db
        self.styles = getSampleStyleSheet()
        self.setup_custom_styles()
    
    def setup_custom_styles(self):
        """Setup custom styles for PDF reports"""
        self.styles.add(ParagraphStyle(
            name='CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            spaceAfter=12,
            textColor=colors.darkblue
        ))
        
        self.styles.add(ParagraphStyle(
            name='CustomSubHeading',
            parent=self.styles['Heading3'],
            fontSize=12,
            spaceAfter=8,
            textColor=colors.darkgreen
        ))

    def generate_report(self, report_type: str, filters: Dict[str, Any], format_type: str = 'pdf') -> bytes:
        """Generate report based on type and format"""
        try:
            # Teacher-level reports
            if report_type == 'student_performance':
                return self.generate_student_performance_report(filters, format_type)
            elif report_type == 'class_analytics':
                return self.generate_class_analytics_report(filters, format_type)
            elif report_type == 'co_po_attainment':
                return self.generate_co_po_attainment_report(filters, format_type)
            elif report_type == 'exam_analysis':
                return self.generate_exam_analysis_report(filters, format_type)
            elif report_type == 'comprehensive_analysis':
                return self.generate_comprehensive_analysis_report(filters, format_type)
            elif report_type == 'custom_analysis':
                return self.generate_custom_analysis_report(filters, format_type)
            
            # CO-PO Analysis Reports
            elif report_type == 'detailed_co_analysis':
                return self.generate_detailed_co_analysis_report(filters, format_type)
            elif report_type == 'detailed_po_analysis':
                return self.generate_detailed_po_analysis_report(filters, format_type)
            elif report_type == 'co_po_mapping_analysis':
                return self.generate_co_po_mapping_analysis_report(filters, format_type)
            elif report_type == 'attainment_gap_analysis':
                return self.generate_attainment_gap_analysis_report(filters, format_type)
            elif report_type == 'comprehensive_co_po_report':
                return self.generate_comprehensive_co_po_report(filters, format_type)
            
            # HOD-level reports
            elif report_type == 'department_performance':
                return self.generate_department_performance_report(filters, format_type)
            elif report_type == 'teacher_effectiveness':
                return self.generate_teacher_effectiveness_report(filters, format_type)
            elif report_type == 'student_progress_tracking':
                return self.generate_student_progress_tracking_report(filters, format_type)
            elif report_type == 'attainment_gap_analysis':
                return self.generate_attainment_gap_analysis_report(filters, format_type)
            elif report_type == 'strategic_planning':
                return self.generate_strategic_planning_report(filters, format_type)
            elif report_type == 'comparative_analysis':
                return self.generate_comparative_analysis_report(filters, format_type)
            elif report_type == 'accreditation_report':
                return self.generate_accreditation_report(filters, format_type)
            elif report_type == 'resource_utilization':
                return self.generate_resource_utilization_report(filters, format_type)
            else:
                raise ValueError(f"Unknown report type: {report_type}")
        except Exception as e:
            raise Exception(f"Error generating report: {str(e)}")

    def generate_student_performance_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate detailed student performance report"""
        subject_id = filters.get('subject_id')
        class_id = filters.get('class_id')
        exam_type = filters.get('exam_type', 'all')
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        if not subject_id:
            raise ValueError("subject_id is required for student performance report")
        
        # Get data
        students = self.get_students_for_class(class_id) if class_id else self.get_all_students()
        subject = self.get_subject(subject_id)
        exams = self.get_exams_for_subject(subject_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_student_performance_pdf(students, subject, exams, filters)
        elif format_type == 'excel':
            return self.generate_student_performance_excel(students, subject, exams, filters)
        elif format_type == 'csv':
            return self.generate_student_performance_csv(students, subject, exams, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_class_analytics_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate comprehensive class analytics report"""
        subject_id = filters.get('subject_id')
        class_id = filters.get('class_id')
        exam_type = filters.get('exam_type', 'all')
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get data
        class_data = self.get_class_analytics_data(subject_id, class_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_class_analytics_pdf(class_data, filters)
        elif format_type == 'excel':
            return self.generate_class_analytics_excel(class_data, filters)
        elif format_type == 'csv':
            return self.generate_class_analytics_csv(class_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_co_po_attainment_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate CO/PO attainment analysis report"""
        subject_id = filters.get('subject_id')
        exam_type = filters.get('exam_type', 'all')
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get data
        co_data = self.get_co_attainment_data(subject_id, exam_type)
        po_data = self.get_po_attainment_data(subject_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_co_po_attainment_pdf(co_data, po_data, filters)
        elif format_type == 'excel':
            return self.generate_co_po_attainment_excel(co_data, po_data, filters)
        elif format_type == 'csv':
            return self.generate_co_po_attainment_csv(co_data, po_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_exam_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate detailed exam analysis report"""
        exam_id = filters.get('exam_id')
        subject_id = filters.get('subject_id')
        
        # Get data
        exam = self.get_exam(exam_id) if exam_id else None
        exam_data = self.get_exam_analysis_data(exam_id or subject_id)
        
        if format_type == 'pdf':
            return self.generate_exam_analysis_pdf(exam, exam_data, filters)
        elif format_type == 'excel':
            return self.generate_exam_analysis_excel(exam, exam_data, filters)
        elif format_type == 'csv':
            return self.generate_exam_analysis_csv(exam, exam_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_comprehensive_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate comprehensive analysis report with all metrics"""
        subject_id = filters.get('subject_id')
        class_id = filters.get('class_id')
        exam_type = filters.get('exam_type', 'all')
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get comprehensive data
        comprehensive_data = self.get_comprehensive_analysis_data(subject_id, class_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_comprehensive_analysis_pdf(comprehensive_data, filters)
        elif format_type == 'excel':
            return self.generate_comprehensive_analysis_excel(comprehensive_data, filters)
        elif format_type == 'csv':
            return self.generate_comprehensive_analysis_csv(comprehensive_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_custom_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate custom analysis report based on specific requirements"""
        # This allows for flexible custom reports based on user requirements
        analysis_type = filters.get('analysis_type', 'general')
        
        if analysis_type == 'performance_trends':
            return self.generate_performance_trends_report(filters, format_type)
        elif analysis_type == 'attainment_gaps':
            return self.generate_attainment_gaps_report(filters, format_type)
        elif analysis_type == 'comparative_analysis':
            return self.generate_comparative_analysis_report(filters, format_type)
        else:
            return self.generate_general_analysis_report(filters, format_type)

    # HOD-specific report generation methods
    def generate_department_performance_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate comprehensive department performance report"""
        department_id = filters.get('department_id')
        academic_year = filters.get('academic_year', '2024-25')
        semester = filters.get('semester', 'all')
        
        if not department_id:
            raise ValueError("department_id is required for department performance report")
        
        # Get department data
        department_data = self.get_department_performance_data(department_id, academic_year, semester)
        
        if format_type == 'pdf':
            return self.generate_department_performance_pdf(department_data, filters)
        elif format_type == 'excel':
            return self.generate_department_performance_excel(department_data, filters)
        elif format_type == 'csv':
            return self.generate_department_performance_csv(department_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_teacher_effectiveness_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate teacher effectiveness analysis report"""
        department_id = filters.get('department_id')
        teacher_id = filters.get('teacher_id')
        academic_year = filters.get('academic_year', '2024-25')
        
        # Get teacher effectiveness data
        teacher_data = self.get_teacher_effectiveness_data(department_id, teacher_id, academic_year)
        
        if format_type == 'pdf':
            return self.generate_teacher_effectiveness_pdf(teacher_data, filters)
        elif format_type == 'excel':
            return self.generate_teacher_effectiveness_excel(teacher_data, filters)
        elif format_type == 'csv':
            return self.generate_teacher_effectiveness_csv(teacher_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_student_progress_tracking_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate student progress tracking report"""
        department_id = filters.get('department_id')
        class_id = filters.get('class_id')
        student_id = filters.get('student_id')
        academic_year = filters.get('academic_year', '2024-25')
        
        # Get student progress data
        progress_data = self.get_student_progress_data(department_id, class_id, student_id, academic_year)
        
        if format_type == 'pdf':
            return self.generate_student_progress_pdf(progress_data, filters)
        elif format_type == 'excel':
            return self.generate_student_progress_excel(progress_data, filters)
        elif format_type == 'csv':
            return self.generate_student_progress_csv(progress_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_attainment_gap_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate attainment gap analysis report"""
        department_id = filters.get('department_id')
        subject_id = filters.get('subject_id')
        academic_year = filters.get('academic_year', '2024-25')
        
        # Get attainment gap data
        gap_data = self.get_attainment_gap_data(department_id, subject_id, academic_year)
        
        if format_type == 'pdf':
            return self.generate_attainment_gap_pdf(gap_data, filters)
        elif format_type == 'excel':
            return self.generate_attainment_gap_excel(gap_data, filters)
        elif format_type == 'csv':
            return self.generate_attainment_gap_csv(gap_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_strategic_planning_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate strategic planning report"""
        department_id = filters.get('department_id')
        academic_year = filters.get('academic_year', '2024-25')
        planning_period = filters.get('planning_period', 'annual')
        
        # Get strategic planning data
        strategic_data = self.get_strategic_planning_data(department_id, academic_year, planning_period)
        
        if format_type == 'pdf':
            return self.generate_strategic_planning_pdf(strategic_data, filters)
        elif format_type == 'excel':
            return self.generate_strategic_planning_excel(strategic_data, filters)
        elif format_type == 'csv':
            return self.generate_strategic_planning_csv(strategic_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_comparative_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate comparative analysis report"""
        department_id = filters.get('department_id')
        comparison_type = filters.get('comparison_type', 'classes')
        time_period = filters.get('time_period', 'current_semester')
        
        # Get comparative analysis data
        comparison_data = self.get_comparative_analysis_data(department_id, comparison_type, time_period)
        
        if format_type == 'pdf':
            return self.generate_comparative_analysis_pdf(comparison_data, filters)
        elif format_type == 'excel':
            return self.generate_comparative_analysis_excel(comparison_data, filters)
        elif format_type == 'csv':
            return self.generate_comparative_analysis_csv(comparison_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_accreditation_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate accreditation report"""
        department_id = filters.get('department_id')
        academic_year = filters.get('academic_year', '2024-25')
        accreditation_standard = filters.get('accreditation_standard', 'NBA')
        
        # Get accreditation data
        accreditation_data = self.get_accreditation_data(department_id, academic_year, accreditation_standard)
        
        if format_type == 'pdf':
            return self.generate_accreditation_pdf(accreditation_data, filters)
        elif format_type == 'excel':
            return self.generate_accreditation_excel(accreditation_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    def generate_resource_utilization_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate resource utilization report"""
        department_id = filters.get('department_id')
        resource_type = filters.get('resource_type', 'all')
        time_period = filters.get('time_period', 'current_semester')
        
        # Get resource utilization data
        resource_data = self.get_resource_utilization_data(department_id, resource_type, time_period)
        
        if format_type == 'pdf':
            return self.generate_resource_utilization_pdf(resource_data, filters)
        elif format_type == 'excel':
            return self.generate_resource_utilization_excel(resource_data, filters)
        elif format_type == 'csv':
            return self.generate_resource_utilization_csv(resource_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")

    # PDF Generation Methods
    def generate_student_performance_pdf(self, students: List, subject: Any, exams: List, filters: Dict) -> bytes:
        """Generate student performance PDF report"""
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        story = []
        
        # Title
        story.append(Paragraph("Student Performance Analysis Report", self.styles['CustomTitle']))
        story.append(Spacer(1, 20))
        
        # Report info
        info_data = [
            ['Subject:', subject.name if subject else 'N/A'],
            ['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
            ['Total Students:', str(len(students))],
            ['Exam Type:', filters.get('exam_type', 'All')]
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 4*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ]))
        story.append(info_table)
        story.append(Spacer(1, 20))
        
        # Student performance table
        story.append(Paragraph("Student Performance Summary", self.styles['CustomHeading']))
        
        # Prepare student data
        student_data = [['Roll No', 'Student Name', 'Total Marks', 'Percentage', 'Grade', 'Status']]
        for student in students:
            performance = self.calculate_student_performance(student.id, exams)
            grade = self.calculate_grade(performance['percentage'])
            status = 'Pass' if performance['percentage'] >= 40 else 'Fail'
            
            student_data.append([
                student.username or '',
                f"{student.first_name} {student.last_name}",
                f"{performance['total_marks']:.2f}",
                f"{performance['percentage']:.2f}%",
                grade,
                status
            ])
        
        student_table = Table(student_data, colWidths=[1*inch, 2*inch, 1*inch, 1*inch, 0.8*inch, 0.8*inch])
        student_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(student_table)
        
        # Statistics summary
        story.append(Spacer(1, 20))
        story.append(Paragraph("Performance Statistics", self.styles['CustomHeading']))
        
        stats = self.calculate_class_statistics(students, exams)
        stats_data = [
            ['Metric', 'Value'],
            ['Average Percentage', f"{stats['average_percentage']:.2f}%"],
            ['Pass Rate', f"{stats['pass_rate']:.2f}%"],
            ['Highest Score', f"{stats['highest_score']:.2f}"],
            ['Lowest Score', f"{stats['lowest_score']:.2f}"],
            ['Standard Deviation', f"{stats['std_deviation']:.2f}"]
        ]
        
        stats_table = Table(stats_data, colWidths=[2*inch, 2*inch])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkgreen),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(stats_table)
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_student_performance_excel(self, students: List, subject: Any, exams: List, filters: Dict) -> bytes:
        """Generate student performance Excel report"""
        buffer = io.BytesIO()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Student Performance"
        
        # Headers
        headers = ['Roll No', 'Student Name', 'Total Marks', 'Percentage', 'Grade', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Student data
        for row, student in enumerate(students, 2):
            performance = self.calculate_student_performance(student.id, exams)
            grade = self.calculate_grade(performance['percentage'])
            status = 'Pass' if performance['percentage'] >= 40 else 'Fail'
            
            ws.cell(row=row, column=1, value=student.username or '')
            ws.cell(row=row, column=2, value=f"{student.first_name} {student.last_name}")
            ws.cell(row=row, column=3, value=performance['total_marks'])
            ws.cell(row=row, column=4, value=performance['percentage'])
            ws.cell(row=row, column=5, value=grade)
            ws.cell(row=row, column=6, value=status)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def generate_student_performance_csv(self, students: List, subject: Any, exams: List, filters: Dict) -> bytes:
        """Generate student performance CSV report"""
        data = []
        for student in students:
            performance = self.calculate_student_performance(student.id, exams)
            grade = self.calculate_grade(performance['percentage'])
            status = 'Pass' if performance['percentage'] >= 40 else 'Fail'
            
            data.append({
                'Roll No': student.username or '',
                'Student Name': f"{student.first_name} {student.last_name}",
                'Total Marks': performance['total_marks'],
                'Percentage': performance['percentage'],
                'Grade': grade,
                'Status': status
            })
        
        df = pd.DataFrame(data)
        return df.to_csv(index=False).encode('utf-8')

    # Data retrieval methods
    def get_students_for_class(self, class_id: int) -> List[User]:
        """Get all students for a specific class"""
        return self.db.query(User).filter(
            and_(User.role == 'student', User.class_id == class_id)
        ).all()

    def get_all_students(self) -> List[User]:
        """Get all students"""
        return self.db.query(User).filter(User.role == 'student').all()

    def get_subject(self, subject_id: int) -> Optional[Subject]:
        """Get subject by ID"""
        return self.db.query(Subject).filter(Subject.id == subject_id).first()

    def get_exams_for_subject(self, subject_id: int, exam_type: str = 'all') -> List[Exam]:
        """Get exams for a subject"""
        query = self.db.query(Exam).filter(Exam.subject_id == subject_id)
        
        # Handle invalid exam types by mapping them to valid ones or 'all'
        valid_exam_types = ['internal1', 'internal2', 'final']
        if exam_type and exam_type != 'all':
            if exam_type in valid_exam_types:
                query = query.filter(Exam.exam_type == exam_type)
            else:
                # Map common invalid types to valid ones
                if exam_type in ['assignment', 'quiz', 'midterm']:
                    query = query.filter(Exam.exam_type == 'internal1')
                # If it's some other invalid type, just return all exams
                
        return query.all()

    def get_exam(self, exam_id: int) -> Optional[Exam]:
        """Get exam by ID"""
        return self.db.query(Exam).filter(Exam.id == exam_id).first()

    def calculate_student_performance(self, student_id: int, exams: List[Exam]) -> Dict[str, float]:
        """Calculate student performance metrics"""
        total_marks = 0
        obtained_marks = 0
        
        for exam in exams:
            marks = self.db.query(Mark).filter(
                and_(Mark.student_id == student_id, Mark.exam_id == exam.id)
            ).all()
            
            exam_total = sum(q.max_marks for q in exam.questions) if exam.questions else 0
            exam_obtained = sum(m.marks_obtained for m in marks)
            
            total_marks += exam_total
            obtained_marks += exam_obtained
        
        percentage = (obtained_marks / total_marks * 100) if total_marks > 0 else 0
        
        return {
            'total_marks': total_marks,
            'obtained_marks': obtained_marks,
            'percentage': percentage
        }

    def calculate_grade(self, percentage: float) -> str:
        """Calculate grade based on percentage"""
        if percentage >= 90:
            return 'A+'
        elif percentage >= 80:
            return 'A'
        elif percentage >= 70:
            return 'B+'
        elif percentage >= 60:
            return 'B'
        elif percentage >= 50:
            return 'C'
        elif percentage >= 40:
            return 'D'
        else:
            return 'F'

    def calculate_class_statistics(self, students: List[User], exams: List[Exam]) -> Dict[str, float]:
        """Calculate class statistics"""
        percentages = []
        for student in students:
            performance = self.calculate_student_performance(student.id, exams)
            percentages.append(performance['percentage'])
        
        if not percentages:
            return {
                'average_percentage': 0,
                'pass_rate': 0,
                'highest_score': 0,
                'lowest_score': 0,
                'std_deviation': 0
            }
        
        import statistics
        pass_count = sum(1 for p in percentages if p >= 40)
        
        return {
            'average_percentage': statistics.mean(percentages),
            'pass_rate': (pass_count / len(percentages)) * 100,
            'highest_score': max(percentages),
            'lowest_score': min(percentages),
            'std_deviation': statistics.stdev(percentages) if len(percentages) > 1 else 0
        }

    # Data retrieval methods
    def get_class_analytics_data(self, subject_id: int, class_id: int, exam_type: str) -> Dict:
        """Get comprehensive class analytics data"""
        subject = self.get_subject(subject_id)
        students = self.get_students_for_class(class_id) if class_id else self.get_all_students()
        exams = self.get_exams_for_subject(subject_id, exam_type)
        
        # Calculate class performance
        class_stats = self.calculate_class_statistics(students, exams)
        
        # Get subject performance
        subject_performance = {
            'subject_name': subject.name if subject else 'Unknown',
            'total_students': len(students),
            'total_exams': len(exams),
            'average_percentage': class_stats['average_percentage'],
            'pass_rate': class_stats['pass_rate'],
            'highest_score': class_stats['highest_score'],
            'lowest_score': class_stats['lowest_score'],
            'std_deviation': class_stats['std_deviation']
        }
        
        return {
            'subject_performance': subject_performance,
            'students': students,
            'exams': exams,
            'class_stats': class_stats
        }
    
    def get_co_attainment_data(self, subject_id: int, exam_type: str) -> Dict:
        """Get comprehensive CO attainment data with detailed analysis"""
        subject = self.get_subject(subject_id)
        if not subject:
            return {}
        
        # Get COs for the subject
        cos = self.db.query(CODefinition).filter(CODefinition.subject_id == subject_id).all()
        exams = self.get_exams_for_subject(subject_id, exam_type)
        
        co_attainment = {}
        co_analysis = {}
        exam_wise_attainment = {}
        
        for co in cos:
            # Calculate attainment for this CO
            total_marks = 0
            obtained_marks = 0
            exam_details = []
            question_analysis = []
            
            for exam in exams:
                # Get questions mapped to this CO through QuestionCOWeight
                questions = self.db.query(Question).join(QuestionCOWeight).filter(
                    and_(Question.exam_id == exam.id, QuestionCOWeight.co_id == co.id)
                ).all()
                
                exam_total = 0
                exam_obtained = 0
                
                for question in questions:
                    # Get marks for this question
                    marks = self.db.query(Mark).filter(Mark.question_id == question.id).all()
                    question_weight = self.db.query(QuestionCOWeight).filter(
                        and_(
                            QuestionCOWeight.question_id == question.id,
                            QuestionCOWeight.co_id == co.id
                        )
                    ).first()
                    
                    if question_weight:
                        question_total = question.max_marks * (question_weight.weight_pct / 100.0)
                        question_obtained = sum(m.marks_obtained for m in marks) * (question_weight.weight_pct / 100.0)
                        
                        total_marks += question_total
                        obtained_marks += question_obtained
                        exam_total += question_total
                        exam_obtained += question_obtained
                        
                        # Question-level analysis
                        question_analysis.append({
                            'question_id': question.id,
                            'question_number': question.question_number,
                            'max_marks': question.max_marks,
                            'weight': question_weight.weight_pct,
                            'weighted_marks': question_total,
                            'obtained_marks': question_obtained,
                            'attainment': (question_obtained / question_total * 100) if question_total > 0 else 0,
                            'exam_name': exam.name,
                            'exam_type': exam.exam_type
                        })
                
                if exam_total > 0:
                    exam_attainment = (exam_obtained / exam_total * 100)
                    exam_details.append({
                        'exam_id': exam.id,
                        'exam_name': exam.name,
                        'exam_type': exam.exam_type,
                        'exam_date': exam.exam_date.isoformat() if exam.exam_date else None,
                        'total_marks': exam_total,
                        'obtained_marks': exam_obtained,
                        'attainment_percentage': exam_attainment,
                        'status': 'Achieved' if exam_attainment >= 70 else 'Not Achieved'
                    })
            
            if total_marks > 0:
                attainment_percentage = (obtained_marks / total_marks * 100)
                
                # Calculate CO analysis metrics
                co_analysis[co.id] = {
                    'attainment_trend': self._calculate_attainment_trend(exam_details),
                    'difficulty_level': self._calculate_difficulty_level(question_analysis),
                    'consistency_score': self._calculate_consistency_score(exam_details),
                    'improvement_areas': self._identify_improvement_areas(question_analysis),
                    'strength_areas': self._identify_strength_areas(question_analysis)
                }
                
                co_attainment[co.id] = {
                    'co_id': co.id,
                    'co_code': co.code,
                    'co_description': co.description,
                    'attainment_percentage': attainment_percentage,
                    'total_marks': total_marks,
                    'obtained_marks': obtained_marks,
                    'target_attainment': 70.0,
                    'status': 'Achieved' if attainment_percentage >= 70 else 'Not Achieved',
                    'gap_analysis': {
                        'gap_percentage': max(0, 70.0 - attainment_percentage),
                        'required_improvement': max(0, (70.0 - attainment_percentage) * total_marks / 100),
                        'priority_level': 'High' if attainment_percentage < 50 else 'Medium' if attainment_percentage < 70 else 'Low'
                    },
                    'exam_wise_attainment': exam_details,
                    'question_analysis': question_analysis,
                    'co_analysis': co_analysis[co.id]
                }
        
        # Calculate overall metrics
        overall_attainment = sum(co['attainment_percentage'] for co in co_attainment.values()) / len(co_attainment) if co_attainment else 0
        achieved_cos = sum(1 for co in co_attainment.values() if co['attainment_percentage'] >= 70)
        total_cos = len(co_attainment)
        
        return {
            'subject_id': subject_id,
            'subject_name': subject.name,
            'subject_code': subject.code,
            'co_attainment': co_attainment,
            'overall_attainment': overall_attainment,
            'summary': {
                'total_cos': total_cos,
                'achieved_cos': achieved_cos,
                'achievement_rate': (achieved_cos / total_cos * 100) if total_cos > 0 else 0,
                'average_attainment': overall_attainment,
                'target_attainment': 70.0,
                'gap_percentage': max(0, 70.0 - overall_attainment),
                'status': 'Target Achieved' if overall_attainment >= 70 else 'Target Not Achieved'
            },
            'recommendations': self._generate_co_recommendations(co_attainment, co_analysis)
        }
    
    def _calculate_attainment_trend(self, exam_details):
        """Calculate attainment trend across exams"""
        if len(exam_details) < 2:
            return 'Insufficient Data'
        
        attainments = [exam['attainment_percentage'] for exam in exam_details]
        if attainments[0] < attainments[-1]:
            return 'Improving'
        elif attainments[0] > attainments[-1]:
            return 'Declining'
        else:
            return 'Stable'
    
    def _calculate_difficulty_level(self, question_analysis):
        """Calculate difficulty level based on question performance"""
        if not question_analysis:
            return 'Unknown'
        
        avg_attainment = sum(q['attainment'] for q in question_analysis) / len(question_analysis)
        if avg_attainment >= 80:
            return 'Easy'
        elif avg_attainment >= 60:
            return 'Medium'
        else:
            return 'Difficult'
    
    def _calculate_consistency_score(self, exam_details):
        """Calculate consistency score across exams"""
        if len(exam_details) < 2:
            return 0
        
        attainments = [exam['attainment_percentage'] for exam in exam_details]
        mean_attainment = sum(attainments) / len(attainments)
        variance = sum((x - mean_attainment) ** 2 for x in attainments) / len(attainments)
        std_dev = variance ** 0.5
        
        # Consistency score (0-100, higher is more consistent)
        consistency = max(0, 100 - (std_dev * 2))
        return round(consistency, 2)
    
    def _identify_improvement_areas(self, question_analysis):
        """Identify areas that need improvement"""
        if not question_analysis:
            return []
        
        low_performance = [q for q in question_analysis if q['attainment'] < 60]
        return [f"Question {q['question_number']}" for q in low_performance[:3]]  # Top 3 areas
    
    def _identify_strength_areas(self, question_analysis):
        """Identify areas of strength"""
        if not question_analysis:
            return []
        
        high_performance = [q for q in question_analysis if q['attainment'] >= 80]
        return [f"Question {q['question_number']}" for q in high_performance[:3]]  # Top 3 areas
    
    def _generate_co_recommendations(self, co_attainment, co_analysis):
        """Generate recommendations based on CO analysis"""
        recommendations = []
        
        for co_id, co_data in co_attainment.items():
            if co_data['attainment_percentage'] < 70:
                priority = co_data['gap_analysis']['priority_level']
                gap = co_data['gap_analysis']['gap_percentage']
                
                if priority == 'High':
                    recommendations.append({
                        'co_code': co_data['co_code'],
                        'priority': 'High',
                        'action': f"Immediate intervention required for {co_data['co_code']}. Gap: {gap:.1f}%",
                        'suggestions': [
                            "Conduct remedial classes",
                            "Review teaching methodology",
                            "Provide additional practice materials"
                        ]
                    })
                elif priority == 'Medium':
                    recommendations.append({
                        'co_code': co_data['co_code'],
                        'priority': 'Medium',
                        'action': f"Focus on improving {co_data['co_code']}. Gap: {gap:.1f}%",
                        'suggestions': [
                            "Increase practice sessions",
                            "Review assessment methods",
                            "Provide targeted feedback"
                        ]
                    })
        
        return recommendations
    
    def get_po_attainment_data(self, subject_id: int, exam_type: str) -> Dict:
        """Get comprehensive PO attainment data with detailed analysis"""
        subject = self.get_subject(subject_id)
        if not subject:
            return {}
        
        # Get POs for the department through the class relationship
        class_obj = self.db.query(Class).filter(Class.id == subject.class_id).first()
        if not class_obj:
            return {}
        
        department_id = class_obj.department_id
        pos = self.db.query(PODefinition).filter(PODefinition.department_id == department_id).all()
        
        po_attainment = {}
        po_analysis = {}
        
        for po in pos:
            # Calculate PO attainment based on CO-PO mapping
            co_po_mappings = self.db.query(COPOMatrix).filter(COPOMatrix.po_id == po.id).all()
            
            total_weight = 0
            weighted_attainment = 0
            co_contributions = []
            
            for mapping in co_po_mappings:
                co = self.db.query(CODefinition).filter(CODefinition.id == mapping.co_id).first()
                if co and co.subject_id == subject_id:
                    co_data = self.get_co_attainment_data(subject_id, exam_type)
                    if co.id in co_data.get('co_attainment', {}):
                        attainment = co_data['co_attainment'][co.id]['attainment_percentage']
                        weight = mapping.weight or 1
                        weighted_attainment += attainment * weight
                        total_weight += weight
                        
                        co_contributions.append({
                            'co_code': co.code,
                            'co_description': co.description,
                            'attainment': attainment,
                            'weight': weight,
                            'contribution': attainment * weight
                        })
            
            po_attainment_percentage = (weighted_attainment / total_weight) if total_weight > 0 else 0
            
            # Calculate PO analysis metrics
            po_analysis[po.id] = {
                'co_contributions': co_contributions,
                'strongest_co': max(co_contributions, key=lambda x: x['contribution']) if co_contributions else None,
                'weakest_co': min(co_contributions, key=lambda x: x['contribution']) if co_contributions else None,
                'contribution_variance': self._calculate_contribution_variance(co_contributions),
                'dependency_analysis': self._analyze_po_dependencies(co_contributions)
            }
            
            po_attainment[po.id] = {
                'po_id': po.id,
                'po_code': po.code,
                'po_description': po.description,
                'attainment_percentage': po_attainment_percentage,
                'target_attainment': 70.0,
                'status': 'Achieved' if po_attainment_percentage >= 70 else 'Not Achieved',
                'gap_analysis': {
                    'gap_percentage': max(0, 70.0 - po_attainment_percentage),
                    'priority_level': 'High' if po_attainment_percentage < 50 else 'Medium' if po_attainment_percentage < 70 else 'Low'
                },
                'co_contributions': co_contributions,
                'po_analysis': po_analysis[po.id]
            }
        
        # Calculate overall metrics
        overall_attainment = sum(po['attainment_percentage'] for po in po_attainment.values()) / len(po_attainment) if po_attainment else 0
        achieved_pos = sum(1 for po in po_attainment.values() if po['attainment_percentage'] >= 70)
        total_pos = len(po_attainment)
        
        return {
            'subject_id': subject_id,
            'subject_name': subject.name,
            'subject_code': subject.code,
            'po_attainment': po_attainment,
            'overall_attainment': overall_attainment,
            'summary': {
                'total_pos': total_pos,
                'achieved_pos': achieved_pos,
                'achievement_rate': (achieved_pos / total_pos * 100) if total_pos > 0 else 0,
                'average_attainment': overall_attainment,
                'target_attainment': 70.0,
                'gap_percentage': max(0, 70.0 - overall_attainment),
                'status': 'Target Achieved' if overall_attainment >= 70 else 'Target Not Achieved'
            },
            'recommendations': self._generate_po_recommendations(po_attainment, po_analysis)
        }
    
    def _calculate_contribution_variance(self, co_contributions):
        """Calculate variance in CO contributions to PO"""
        if len(co_contributions) < 2:
            return 0
        
        contributions = [c['contribution'] for c in co_contributions]
        mean_contrib = sum(contributions) / len(contributions)
        variance = sum((x - mean_contrib) ** 2 for x in contributions) / len(contributions)
        return round(variance, 2)
    
    def _analyze_po_dependencies(self, co_contributions):
        """Analyze PO dependencies on COs"""
        if not co_contributions:
            return {'dependency_level': 'Unknown', 'critical_cos': []}
        
        # Sort COs by contribution
        sorted_cos = sorted(co_contributions, key=lambda x: x['contribution'], reverse=True)
        
        # Identify critical COs (top 50% contributors)
        critical_count = max(1, len(sorted_cos) // 2)
        critical_cos = sorted_cos[:critical_count]
        
        # Calculate dependency level
        total_contribution = sum(c['contribution'] for c in co_contributions)
        critical_contribution = sum(c['contribution'] for c in critical_cos)
        dependency_ratio = critical_contribution / total_contribution if total_contribution > 0 else 0
        
        if dependency_ratio > 0.8:
            dependency_level = 'High'
        elif dependency_ratio > 0.6:
            dependency_level = 'Medium'
        else:
            dependency_level = 'Low'
        
        return {
            'dependency_level': dependency_level,
            'critical_cos': [c['co_code'] for c in critical_cos],
            'dependency_ratio': round(dependency_ratio, 2)
        }
    
    def _generate_po_recommendations(self, po_attainment, po_analysis):
        """Generate recommendations based on PO analysis"""
        recommendations = []
        
        for po_id, po_data in po_attainment.items():
            if po_data['attainment_percentage'] < 70:
                priority = po_data['gap_analysis']['priority_level']
                gap = po_data['gap_analysis']['gap_percentage']
                
                # Get critical COs for this PO
                critical_cos = po_analysis[po_id]['dependency_analysis']['critical_cos']
                
                if priority == 'High':
                    recommendations.append({
                        'po_code': po_data['po_code'],
                        'priority': 'High',
                        'action': f"Immediate intervention required for {po_data['po_code']}. Gap: {gap:.1f}%",
                        'critical_cos': critical_cos,
                        'suggestions': [
                            f"Focus on improving critical COs: {', '.join(critical_cos)}",
                            "Review CO-PO mapping weights",
                            "Conduct comprehensive assessment"
                        ]
                    })
                elif priority == 'Medium':
                    recommendations.append({
                        'po_code': po_data['po_code'],
                        'priority': 'Medium',
                        'action': f"Focus on improving {po_data['po_code']}. Gap: {gap:.1f}%",
                        'critical_cos': critical_cos,
                        'suggestions': [
                            f"Strengthen critical COs: {', '.join(critical_cos)}",
                            "Review assessment strategies",
                            "Provide targeted support"
                        ]
                    })
        
        return recommendations
    
    def generate_detailed_co_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate detailed CO analysis report"""
        subject_id = filters.get('subject_id')
        exam_type = filters.get('exam_type', 'all')
        
        if not subject_id:
            raise ValueError("subject_id is required for detailed CO analysis report")
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get comprehensive CO data
        co_data = self.get_co_attainment_data(subject_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_detailed_co_analysis_pdf(co_data, filters)
        elif format_type == 'excel':
            return self.generate_detailed_co_analysis_excel(co_data, filters)
        elif format_type == 'csv':
            return self.generate_detailed_co_analysis_csv(co_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    def generate_detailed_co_analysis_pdf(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed CO analysis PDF report"""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=20,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Detailed Course Outcome (CO) Analysis Report", title_style))
        story.append(Spacer(1, 12))
        
        # Subject info
        story.append(Paragraph(f"<b>Subject:</b> {data.get('subject_name', 'N/A')} ({data.get('subject_code', 'N/A')})", styles['Normal']))
        story.append(Paragraph(f"<b>Exam Type:</b> {filters.get('exam_type', 'All')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Executive Summary
        summary = data.get('summary', {})
        story.append(Paragraph("<b>Executive Summary</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        summary_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Total COs', str(summary.get('total_cos', 0)), 'N/A', 'Active'],
            ['Achieved COs', str(summary.get('achieved_cos', 0)), 'N/A', 'Completed'],
            ['Achievement Rate', f"{summary.get('achievement_rate', 0):.1f}%", '100%', 'Good' if summary.get('achievement_rate', 0) >= 80 else 'Needs Improvement'],
            ['Average Attainment', f"{summary.get('average_attainment', 0):.1f}%", '70%', 'Good' if summary.get('average_attainment', 0) >= 70 else 'Needs Improvement'],
            ['Gap Percentage', f"{summary.get('gap_percentage', 0):.1f}%", '0%', 'High' if summary.get('gap_percentage', 0) > 20 else 'Medium' if summary.get('gap_percentage', 0) > 10 else 'Low'],
        ]
        
        summary_table = Table(summary_data)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 20))
        
        # Detailed CO Analysis
        story.append(Paragraph("<b>Detailed CO Analysis</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        co_attainment = data.get('co_attainment', {})
        for co_id, co_info in co_attainment.items():
            # CO Header
            story.append(Paragraph(f"<b>{co_info.get('co_code', 'N/A')}: {co_info.get('co_description', 'N/A')}</b>", styles['Heading3']))
            story.append(Spacer(1, 8))
            
            # CO Metrics
            co_metrics = [
                ['Metric', 'Value', 'Target', 'Status'],
                ['Attainment %', f"{co_info.get('attainment_percentage', 0):.1f}%", '70%', co_info.get('status', 'N/A')],
                ['Total Marks', f"{co_info.get('total_marks', 0):.1f}", 'N/A', 'N/A'],
                ['Obtained Marks', f"{co_info.get('obtained_marks', 0):.1f}", 'N/A', 'N/A'],
                ['Gap %', f"{co_info.get('gap_analysis', {}).get('gap_percentage', 0):.1f}%", '0%', co_info.get('gap_analysis', {}).get('priority_level', 'N/A')],
            ]
            
            co_table = Table(co_metrics)
            co_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(co_table)
            story.append(Spacer(1, 12))
            
            # CO Analysis Details
            co_analysis = co_info.get('co_analysis', {})
            if co_analysis:
                story.append(Paragraph("<b>Analysis Details:</b>", styles['Normal']))
                analysis_text = f"""
                • Attainment Trend: {co_analysis.get('attainment_trend', 'N/A')}<br/>
                • Difficulty Level: {co_analysis.get('difficulty_level', 'N/A')}<br/>
                • Consistency Score: {co_analysis.get('consistency_score', 0)}/100<br/>
                • Strength Areas: {', '.join(co_analysis.get('strength_areas', [])[:2]) or 'None identified'}<br/>
                • Improvement Areas: {', '.join(co_analysis.get('improvement_areas', [])[:2]) or 'None identified'}
                """
                story.append(Paragraph(analysis_text, styles['Normal']))
                story.append(Spacer(1, 12))
            
            # Exam-wise Performance
            exam_details = co_info.get('exam_wise_attainment', [])
            if exam_details:
                story.append(Paragraph("<b>Exam-wise Performance:</b>", styles['Normal']))
                exam_data = [['Exam', 'Type', 'Date', 'Attainment %', 'Status']]
                for exam in exam_details:
                    exam_data.append([
                        exam.get('exam_name', 'N/A'),
                        exam.get('exam_type', 'N/A'),
                        exam.get('exam_date', 'N/A')[:10] if exam.get('exam_date') else 'N/A',
                        f"{exam.get('attainment_percentage', 0):.1f}%",
                        exam.get('status', 'N/A')
                    ])
                
                exam_table = Table(exam_data)
                exam_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.white),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(exam_table)
                story.append(Spacer(1, 12))
            
            story.append(PageBreak())
        
        # Recommendations
        recommendations = data.get('recommendations', [])
        if recommendations:
            story.append(Paragraph("<b>Recommendations</b>", styles['Heading2']))
            story.append(Spacer(1, 12))
            
            for i, rec in enumerate(recommendations, 1):
                story.append(Paragraph(f"<b>{i}. {rec.get('co_code', 'N/A')} - {rec.get('priority', 'N/A')} Priority</b>", styles['Normal']))
                story.append(Paragraph(f"Action: {rec.get('action', 'N/A')}", styles['Normal']))
                story.append(Paragraph("Suggestions:", styles['Normal']))
                for suggestion in rec.get('suggestions', []):
                    story.append(Paragraph(f"• {suggestion}", styles['Normal']))
                story.append(Spacer(1, 8))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_detailed_co_analysis_excel(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed CO analysis Excel report"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        wb = openpyxl.Workbook()
        
        # Summary Sheet
        ws_summary = wb.active
        ws_summary.title = "CO Analysis Summary"
        
        # Title
        ws_summary['A1'] = "Detailed Course Outcome (CO) Analysis Report"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:E1')
        
        # Subject info
        ws_summary['A3'] = f"Subject: {data.get('subject_name', 'N/A')} ({data.get('subject_code', 'N/A')})"
        ws_summary['A4'] = f"Exam Type: {filters.get('exam_type', 'All')}"
        
        # Summary metrics
        summary = data.get('summary', {})
        ws_summary['A6'] = "Executive Summary"
        ws_summary['A6'].font = Font(size=14, bold=True)
        
        summary_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Total COs', summary.get('total_cos', 0), 'N/A', 'Active'],
            ['Achieved COs', summary.get('achieved_cos', 0), 'N/A', 'Completed'],
            ['Achievement Rate', f"{summary.get('achievement_rate', 0):.1f}%", '100%', 'Good' if summary.get('achievement_rate', 0) >= 80 else 'Needs Improvement'],
            ['Average Attainment', f"{summary.get('average_attainment', 0):.1f}%", '70%', 'Good' if summary.get('average_attainment', 0) >= 70 else 'Needs Improvement'],
            ['Gap Percentage', f"{summary.get('gap_percentage', 0):.1f}%", '0%', 'High' if summary.get('gap_percentage', 0) > 20 else 'Medium' if summary.get('gap_percentage', 0) > 10 else 'Low'],
        ]
        
        for row_num, row_data in enumerate(summary_data, 7):
            for col_num, value in enumerate(row_data, 1):
                cell = ws_summary.cell(row=row_num, column=col_num, value=value)
                if row_num == 7:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Detailed CO Analysis Sheet
        ws_details = wb.create_sheet("Detailed CO Analysis")
        
        # Headers
        headers = ['CO Code', 'CO Description', 'Attainment %', 'Status', 'Gap %', 'Priority', 'Trend', 'Difficulty', 'Consistency']
        for col_num, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # CO Data
        co_attainment = data.get('co_attainment', {})
        row_num = 2
        for co_id, co_info in co_attainment.items():
            co_analysis = co_info.get('co_analysis', {})
            ws_details.cell(row=row_num, column=1, value=co_info.get('co_code', 'N/A'))
            ws_details.cell(row=row_num, column=2, value=co_info.get('co_description', 'N/A'))
            ws_details.cell(row=row_num, column=3, value=f"{co_info.get('attainment_percentage', 0):.1f}%")
            ws_details.cell(row=row_num, column=4, value=co_info.get('status', 'N/A'))
            ws_details.cell(row=row_num, column=5, value=f"{co_info.get('gap_analysis', {}).get('gap_percentage', 0):.1f}%")
            ws_details.cell(row=row_num, column=6, value=co_info.get('gap_analysis', {}).get('priority_level', 'N/A'))
            ws_details.cell(row=row_num, column=7, value=co_analysis.get('attainment_trend', 'N/A'))
            ws_details.cell(row=row_num, column=8, value=co_analysis.get('difficulty_level', 'N/A'))
            ws_details.cell(row=row_num, column=9, value=f"{co_analysis.get('consistency_score', 0)}/100")
            row_num += 1
        
        # Auto-adjust column widths
        for ws in [ws_summary, ws_details]:
            for column in ws.columns:
                max_length = 0
                column_letter = None
                for cell in column:
                    try:
                        if hasattr(cell, 'column_letter'):
                            if column_letter is None:
                                column_letter = cell.column_letter
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                    except:
                        pass
                if column_letter:
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_detailed_co_analysis_csv(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed CO analysis CSV report"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Detailed Course Outcome (CO) Analysis Report"])
        writer.writerow([])
        writer.writerow([f"Subject: {data.get('subject_name', 'N/A')} ({data.get('subject_code', 'N/A')})"])
        writer.writerow([f"Exam Type: {filters.get('exam_type', 'All')}"])
        writer.writerow([])
        
        # Summary
        summary = data.get('summary', {})
        writer.writerow(["Executive Summary"])
        writer.writerow(["Metric", "Value", "Target", "Status"])
        writer.writerow(["Total COs", summary.get('total_cos', 0), "N/A", "Active"])
        writer.writerow(["Achieved COs", summary.get('achieved_cos', 0), "N/A", "Completed"])
        writer.writerow(["Achievement Rate", f"{summary.get('achievement_rate', 0):.1f}%", "100%", "Good" if summary.get('achievement_rate', 0) >= 80 else "Needs Improvement"])
        writer.writerow(["Average Attainment", f"{summary.get('average_attainment', 0):.1f}%", "70%", "Good" if summary.get('average_attainment', 0) >= 70 else "Needs Improvement"])
        writer.writerow(["Gap Percentage", f"{summary.get('gap_percentage', 0):.1f}%", "0%", "High" if summary.get('gap_percentage', 0) > 20 else "Medium" if summary.get('gap_percentage', 0) > 10 else "Low"])
        writer.writerow([])
        
        # Detailed CO Analysis
        writer.writerow(["Detailed CO Analysis"])
        writer.writerow(["CO Code", "CO Description", "Attainment %", "Status", "Gap %", "Priority", "Trend", "Difficulty", "Consistency"])
        
        co_attainment = data.get('co_attainment', {})
        for co_id, co_info in co_attainment.items():
            co_analysis = co_info.get('co_analysis', {})
            writer.writerow([
                co_info.get('co_code', 'N/A'),
                co_info.get('co_description', 'N/A'),
                f"{co_info.get('attainment_percentage', 0):.1f}%",
                co_info.get('status', 'N/A'),
                f"{co_info.get('gap_analysis', {}).get('gap_percentage', 0):.1f}%",
                co_info.get('gap_analysis', {}).get('priority_level', 'N/A'),
                co_analysis.get('attainment_trend', 'N/A'),
                co_analysis.get('difficulty_level', 'N/A'),
                f"{co_analysis.get('consistency_score', 0)}/100"
            ])
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def generate_detailed_po_analysis_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate detailed PO analysis report"""
        subject_id = filters.get('subject_id')
        exam_type = filters.get('exam_type', 'all')
        
        if not subject_id:
            raise ValueError("subject_id is required for detailed PO analysis report")
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get comprehensive PO data
        po_data = self.get_po_attainment_data(subject_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_detailed_po_analysis_pdf(po_data, filters)
        elif format_type == 'excel':
            return self.generate_detailed_po_analysis_excel(po_data, filters)
        elif format_type == 'csv':
            return self.generate_detailed_po_analysis_csv(po_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    def generate_detailed_po_analysis_pdf(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed PO analysis PDF report"""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Detailed PO Analysis Report", title_style))
        story.append(Spacer(1, 12))
        
        # Summary
        if data.get('summary'):
            story.append(Paragraph("Summary", styles['Heading2']))
            summary_text = data['summary'] if isinstance(data['summary'], str) else str(data['summary'])
            story.append(Paragraph(summary_text, styles['Normal']))
            story.append(Spacer(1, 12))
        
        # PO Attainment Data
        if data.get('po_attainment'):
            story.append(Paragraph("PO Attainment Analysis", styles['Heading2']))
            
            # Create table for PO data
            po_data = []
            po_data.append(['PO Code', 'PO Title', 'Attainment %', 'Status'])
            
            for po_id, po_info in data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                po_data.append([
                    po_info.get('po_code', 'N/A'),
                    po_info.get('po_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status
                ])
            
            po_table = Table(po_data)
            po_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(po_table)
            story.append(Spacer(1, 12))
        
        # Recommendations
        if data.get('recommendations'):
            story.append(Paragraph("Recommendations", styles['Heading2']))
            for i, rec in enumerate(data['recommendations'], 1):
                story.append(Paragraph(f"{i}. {rec}", styles['Normal']))
        
        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()
        return pdf_content
    
    def generate_detailed_po_analysis_excel(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed PO analysis Excel report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "PO Analysis"
        
        # Headers
        headers = ['PO Code', 'PO Title', 'Attainment %', 'Status', 'CO Contributions']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Data
        row = 2
        if data.get('po_attainment'):
            for po_id, po_info in data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                co_contributions = ', '.join([co.get('co_code', '') for co in po_info.get('co_contributions', [])])
                
                ws.cell(row=row, column=1, value=po_info.get('po_code', 'N/A'))
                ws.cell(row=row, column=2, value=po_info.get('po_title', 'N/A'))
                ws.cell(row=row, column=3, value=attainment)
                ws.cell(row=row, column=4, value=status)
                ws.cell(row=row, column=5, value=co_contributions)
                row += 1
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        return excel_content
    
    def generate_detailed_po_analysis_csv(self, data: Dict, filters: Dict) -> bytes:
        """Generate detailed PO analysis CSV report"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Headers
        writer.writerow(['PO Code', 'PO Title', 'Attainment %', 'Status', 'CO Contributions'])
        
        # Data
        if data.get('po_attainment'):
            for po_id, po_info in data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                co_contributions = ', '.join([co.get('co_code', '') for co in po_info.get('co_contributions', [])])
                
                writer.writerow([
                    po_info.get('po_code', 'N/A'),
                    po_info.get('po_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status,
                    co_contributions
                ])
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def generate_comprehensive_co_po_report(self, filters: Dict[str, Any], format_type: str) -> bytes:
        """Generate comprehensive CO-PO analysis report"""
        subject_id = filters.get('subject_id')
        exam_type = filters.get('exam_type', 'all')
        
        if not subject_id:
            raise ValueError("subject_id is required for comprehensive CO-PO report")
        
        # Handle empty string exam_type
        if not exam_type or exam_type.strip() == '':
            exam_type = 'all'
        
        # Get both CO and PO data
        co_data = self.get_co_attainment_data(subject_id, exam_type)
        po_data = self.get_po_attainment_data(subject_id, exam_type)
        
        if format_type == 'pdf':
            return self.generate_comprehensive_co_po_pdf(co_data, po_data, filters)
        elif format_type == 'excel':
            return self.generate_comprehensive_co_po_excel(co_data, po_data, filters)
        elif format_type == 'csv':
            return self.generate_comprehensive_co_po_csv(co_data, po_data, filters)
        else:
            raise ValueError(f"Unsupported format: {format_type}")
    
    def generate_comprehensive_co_po_pdf(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        """Generate comprehensive CO-PO PDF report"""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Comprehensive CO-PO Analysis Report", title_style))
        story.append(Spacer(1, 12))
        
        # CO Analysis Section
        story.append(Paragraph("Course Outcomes (CO) Analysis", styles['Heading2']))
        if co_data.get('co_attainment'):
            co_table_data = [['CO Code', 'CO Title', 'Attainment %', 'Status']]
            for co_id, co_info in co_data['co_attainment'].items():
                attainment = co_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                co_table_data.append([
                    co_info.get('co_code', 'N/A'),
                    co_info.get('co_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status
                ])
            
            co_table = Table(co_table_data)
            co_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(co_table)
        
        story.append(PageBreak())
        
        # PO Analysis Section
        story.append(Paragraph("Program Outcomes (PO) Analysis", styles['Heading2']))
        if po_data.get('po_attainment'):
            po_table_data = [['PO Code', 'PO Title', 'Attainment %', 'Status']]
            for po_id, po_info in po_data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                po_table_data.append([
                    po_info.get('po_code', 'N/A'),
                    po_info.get('po_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status
                ])
            
            po_table = Table(po_table_data)
            po_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 14),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(po_table)
        
        doc.build(story)
        pdf_content = buffer.getvalue()
        buffer.close()
        return pdf_content
    
    def generate_comprehensive_co_po_excel(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        """Generate comprehensive CO-PO Excel report"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = Workbook()
        
        # CO Analysis Sheet
        co_ws = wb.active
        co_ws.title = "CO Analysis"
        
        co_headers = ['CO Code', 'CO Title', 'Attainment %', 'Status']
        for col, header in enumerate(co_headers, 1):
            cell = co_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        if co_data.get('co_attainment'):
            for co_id, co_info in co_data['co_attainment'].items():
                attainment = co_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                
                co_ws.cell(row=row, column=1, value=co_info.get('co_code', 'N/A'))
                co_ws.cell(row=row, column=2, value=co_info.get('co_title', 'N/A'))
                co_ws.cell(row=row, column=3, value=attainment)
                co_ws.cell(row=row, column=4, value=status)
                row += 1
        
        # PO Analysis Sheet
        po_ws = wb.create_sheet("PO Analysis")
        
        po_headers = ['PO Code', 'PO Title', 'Attainment %', 'Status']
        for col, header in enumerate(po_headers, 1):
            cell = po_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        row = 2
        if po_data.get('po_attainment'):
            for po_id, po_info in po_data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                
                po_ws.cell(row=row, column=1, value=po_info.get('po_code', 'N/A'))
                po_ws.cell(row=row, column=2, value=po_info.get('po_title', 'N/A'))
                po_ws.cell(row=row, column=3, value=attainment)
                po_ws.cell(row=row, column=4, value=status)
                row += 1
        
        # Auto-adjust column widths
        for ws in [co_ws, po_ws]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        excel_content = buffer.getvalue()
        buffer.close()
        return excel_content
    
    def generate_comprehensive_co_po_csv(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        """Generate comprehensive CO-PO CSV report"""
        import csv
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # CO Analysis
        writer.writerow(['CO Analysis'])
        writer.writerow(['CO Code', 'CO Title', 'Attainment %', 'Status'])
        
        if co_data.get('co_attainment'):
            for co_id, co_info in co_data['co_attainment'].items():
                attainment = co_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                
                writer.writerow([
                    co_info.get('co_code', 'N/A'),
                    co_info.get('co_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status
                ])
        
        writer.writerow([])  # Empty row
        
        # PO Analysis
        writer.writerow(['PO Analysis'])
        writer.writerow(['PO Code', 'PO Title', 'Attainment %', 'Status'])
        
        if po_data.get('po_attainment'):
            for po_id, po_info in po_data['po_attainment'].items():
                attainment = po_info.get('attainment_percentage', 0)
                status = 'Achieved' if attainment >= 70 else 'Not Achieved'
                
                writer.writerow([
                    po_info.get('po_code', 'N/A'),
                    po_info.get('po_title', 'N/A'),
                    f"{attainment:.1f}%",
                    status
                ])
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def get_co_po_mapping_data(self, subject_id: int) -> List[Dict]:
        """Get CO-PO mapping data for a subject"""
        # Get CO-PO mappings for the subject
        mappings = self.db.query(COPOMatrix).filter(COPOMatrix.subject_id == subject_id).all()
        
        mapping_data = []
        for mapping in mappings:
            # Get CO details
            co = self.db.query(CODefinition).filter(CODefinition.id == mapping.co_id).first()
            # Get PO details
            po = self.db.query(PODefinition).filter(PODefinition.id == mapping.po_id).first()
            
            if co and po:
                mapping_data.append({
                    'id': mapping.id,
                    'co_id': mapping.co_id,
                    'po_id': mapping.po_id,
                    'co_code': mapping.co_code,
                    'po_code': mapping.po_code,
                    'strength': mapping.strength,
                    'co_description': co.description,
                    'po_description': po.description
                })
        
        return mapping_data
    
    def get_exam_analysis_data(self, exam_id: int) -> Dict:
        """Get exam analysis data"""
        exam = self.get_exam(exam_id)
        if not exam:
            return {}
        
        # Get all marks for this exam
        marks = self.db.query(Mark).filter(Mark.exam_id == exam_id).all()
        
        # Calculate exam statistics
        total_possible = sum(q.max_marks for q in exam.questions) if exam.questions else 0
        student_scores = {}
        
        for mark in marks:
            if mark.student_id not in student_scores:
                student_scores[mark.student_id] = 0
            student_scores[mark.student_id] += mark.marks_obtained
        
        percentages = [(score / total_possible * 100) if total_possible > 0 else 0 for score in student_scores.values()]
        
        import statistics
        exam_stats = {
            'exam_name': exam.name,
            'exam_type': exam.exam_type,
            'total_marks': total_possible,
            'total_students': len(student_scores),
            'average_percentage': statistics.mean(percentages) if percentages else 0,
            'highest_percentage': max(percentages) if percentages else 0,
            'lowest_percentage': min(percentages) if percentages else 0,
            'pass_rate': (sum(1 for p in percentages if p >= 40) / len(percentages) * 100) if percentages else 0,
            'std_deviation': statistics.stdev(percentages) if len(percentages) > 1 else 0
        }
        
        return {
            'exam': exam,
            'exam_stats': exam_stats,
            'student_scores': student_scores
        }
    
    def get_comprehensive_analysis_data(self, subject_id: int, class_id: int, exam_type: str) -> Dict:
        """Get comprehensive analysis data"""
        # Get all the data
        class_data = self.get_class_analytics_data(subject_id, class_id, exam_type)
        co_data = self.get_co_attainment_data(subject_id, exam_type)
        po_data = self.get_po_attainment_data(subject_id, exam_type)
        
        return {
            'class_analytics': class_data,
            'co_attainment': co_data,
            'po_attainment': po_data,
            'summary': {
                'subject_name': class_data.get('subject_performance', {}).get('subject_name', 'Unknown'),
                'total_students': class_data.get('subject_performance', {}).get('total_students', 0),
                'overall_performance': class_data.get('class_stats', {}).get('average_percentage', 0),
                'co_attainment': co_data.get('overall_attainment', 0),
                'po_attainment': po_data.get('overall_attainment', 0)
            }
        }

    # HOD-specific data retrieval methods
    def get_department_performance_data(self, department_id: int, academic_year: str, semester: str) -> Dict:
        """Get comprehensive department performance data"""
        # Get all subjects in the department
        subjects = self.db.query(Subject).join(Class).filter(Class.department_id == department_id).all()
        
        # Get department info
        department = self.db.query(Department).filter(Department.id == department_id).first()
        
        department_stats = {
            'department': {
                'id': department_id,
                'name': department.name if department else 'Unknown Department',
                'academic_year': academic_year,
                'semester': semester
            },
            'total_students': 0,
            'avg_performance': 0,
            'co_attainment': 0,
            'po_attainment': 0,
            'subject_performance': [],
            'total_subjects': len(subjects),
            'overall_performance': 0,
            'class_performance': {},
            'attainment_summary': {}
        }
        
        total_performance = 0
        subject_count = 0
        
        total_co_attainment = 0
        total_po_attainment = 0
        
        for subject in subjects:
            # Get class performance for this subject
            class_data = self.get_class_analytics_data(subject.id, None, 'all')
            co_data = self.get_co_attainment_data(subject.id, 'all')
            
            subject_perf = {
                'name': subject.name,
                'code': subject.code,
                'student_count': class_data.get('subject_performance', {}).get('total_students', 0),
                'avg_score': class_data.get('class_stats', {}).get('average_percentage', 0),
                'pass_rate': class_data.get('class_stats', {}).get('pass_rate', 0),
                'co_attainment': co_data.get('overall_attainment', 0)
            }
            
            department_stats['subject_performance'].append(subject_perf)
            department_stats['total_students'] += subject_perf['student_count']
            total_performance += subject_perf['avg_score']
            total_co_attainment += subject_perf['co_attainment']
            subject_count += 1
        
        if subject_count > 0:
            department_stats['avg_performance'] = total_performance / subject_count
            department_stats['co_attainment'] = total_co_attainment / subject_count
            department_stats['po_attainment'] = total_co_attainment / subject_count  # Using CO as proxy for PO
            department_stats['overall_performance'] = total_performance / subject_count
        
        return department_stats

    def get_teacher_effectiveness_data(self, department_id: int, teacher_id: int, academic_year: str) -> Dict:
        """Get teacher effectiveness data"""
        # Get teachers in the department
        if teacher_id:
            teachers = [self.db.query(User).filter(User.id == teacher_id).first()]
        else:
            teachers = self.db.query(User).filter(
                and_(User.role == 'teacher', User.department_id == department_id)
            ).all()
        
        # For single teacher report, return the first teacher's data
        if teachers and teachers[0]:
            teacher = teachers[0]
            # Get subjects taught by this teacher
            subjects = self.db.query(Subject).filter(Subject.teacher_id == teacher.id).all()
            
            teacher_data = {
                'teacher': {
                    'id': teacher.id,
                    'name': f"{teacher.first_name} {teacher.last_name}",
                    'department': 'Department Name'  # You can get this from department table
                },
                'overall_rating': 4.2,  # Mock data - replace with actual calculation
                'student_satisfaction': 85.5,  # Mock data - replace with actual calculation
                'class_performance': 78.3,  # Mock data - replace with actual calculation
                'co_attainment': 82.1,  # Mock data - replace with actual calculation
                'subject_performance': []
            }
            
            total_performance = 0
            subject_count = 0
            
            for subject in subjects:
                class_data = self.get_class_analytics_data(subject.id, None, 'all')
                co_data = self.get_co_attainment_data(subject.id, 'all')
                
                subject_perf = {
                    'name': subject.name,
                    'code': subject.code,
                    'avg_score': class_data.get('class_stats', {}).get('average_percentage', 0),
                    'student_count': class_data.get('subject_performance', {}).get('total_students', 0),
                    'rating': 4.0 + (class_data.get('class_stats', {}).get('average_percentage', 0) - 70) / 10  # Mock rating calculation
                }
                
                teacher_data['subject_performance'].append(subject_perf)
                total_performance += subject_perf['avg_score']
                subject_count += 1
            
            if subject_count > 0:
                teacher_data['class_performance'] = total_performance / subject_count
                teacher_data['overall_rating'] = 4.0 + (teacher_data['class_performance'] - 70) / 10
                teacher_data['student_satisfaction'] = min(100, 70 + (teacher_data['class_performance'] - 70) * 0.5)
                teacher_data['co_attainment'] = min(100, 70 + (teacher_data['class_performance'] - 70) * 0.3)
            
            return teacher_data
        
        # Return empty data if no teachers found
        return {
            'teacher': {
                'id': 0,
                'name': 'No Teacher Found',
                'department': 'Unknown'
            },
            'overall_rating': 0,
            'student_satisfaction': 0,
            'class_performance': 0,
            'co_attainment': 0,
            'subject_performance': []
        }

    def get_student_progress_data(self, department_id: int, class_id: int, student_id: int, academic_year: str) -> Dict:
        """Get student progress tracking data"""
        # Get students based on filters
        if student_id:
            students = [self.db.query(User).filter(User.id == student_id).first()]
        elif class_id:
            students = self.get_students_for_class(class_id)
        else:
            # Get all students in the department
            students = self.db.query(User).join(Class).filter(
                and_(User.role == 'student', Class.department_id == department_id)
            ).all()
        
        progress_data = {}
        
        for student in students:
            if not student:
                continue
                
            # Get all subjects for this student
            subjects = self.db.query(Subject).join(Class).filter(
                and_(Class.id == student.class_id, Class.department_id == department_id)
            ).all()
            
            student_progress = {
                'student_id': student.id,
                'student_name': f"{student.first_name} {student.last_name}",
                'roll_number': student.username,
                'class_name': student.class_id,
                'total_subjects': len(subjects),
                'subject_progress': {},
                'overall_progress': 0,
                'grade_point_average': 0
            }
            
            total_performance = 0
            subject_count = 0
            
            for subject in subjects:
                # Get student performance in this subject
                exams = self.get_exams_for_subject(subject.id, 'all')
                performance = self.calculate_student_performance(student.id, exams)
                
                subject_progress = {
                    'subject_name': subject.name,
                    'subject_code': subject.code,
                    'total_marks': performance['total_marks'],
                    'obtained_marks': performance['obtained_marks'],
                    'percentage': performance['percentage'],
                    'grade': self.calculate_grade(performance['percentage']),
                    'status': 'Pass' if performance['percentage'] >= 40 else 'Fail'
                }
                
                student_progress['subject_progress'][subject.id] = subject_progress
                total_performance += performance['percentage']
                subject_count += 1
            
            if subject_count > 0:
                student_progress['overall_progress'] = total_performance / subject_count
                # Calculate GPA (simplified)
                student_progress['grade_point_average'] = self.calculate_gpa(total_performance / subject_count)
            
            progress_data[student.id] = student_progress
        
        return {
            'department_id': department_id,
            'academic_year': academic_year,
            'students': progress_data,
            'summary': {
                'total_students': len(progress_data),
                'average_progress': sum(s['overall_progress'] for s in progress_data.values()) / len(progress_data) if progress_data else 0
            }
        }

    def get_attainment_gap_data(self, department_id: int, subject_id: int, academic_year: str) -> Dict:
        """Get attainment gap analysis data"""
        if subject_id:
            subjects = [self.get_subject(subject_id)]
        else:
            subjects = self.db.query(Subject).join(Class).filter(Class.department_id == department_id).all()
        
        gap_analysis = {}
        
        for subject in subjects:
            if not subject:
                continue
                
            co_data = self.get_co_attainment_data(subject.id, 'all')
            po_data = self.get_po_attainment_data(subject.id, 'all')
            
            gaps = []
            for co_id, co_info in co_data.get('co_attainment', {}).items():
                if co_info['attainment_percentage'] < co_info['target_attainment']:
                    gaps.append({
                        'co_code': co_info['co_code'],
                        'current_attainment': co_info['attainment_percentage'],
                        'target_attainment': co_info['target_attainment'],
                        'gap': co_info['target_attainment'] - co_info['attainment_percentage'],
                        'priority': 'High' if co_info['attainment_percentage'] < 50 else 'Medium'
                    })
            
            gap_analysis[subject.id] = {
                'subject_name': subject.name,
                'subject_code': subject.code,
                'co_attainment': co_data.get('overall_attainment', 0),
                'po_attainment': po_data.get('overall_attainment', 0),
                'gaps': gaps,
                'total_gaps': len(gaps),
                'critical_gaps': len([g for g in gaps if g['priority'] == 'High'])
            }
        
        return {
            'department_id': department_id,
            'academic_year': academic_year,
            'subjects': gap_analysis,
            'summary': {
                'total_subjects': len(gap_analysis),
                'total_gaps': sum(s['total_gaps'] for s in gap_analysis.values()),
                'critical_gaps': sum(s['critical_gaps'] for s in gap_analysis.values())
            }
        }

    def get_strategic_planning_data(self, department_id: int, academic_year: str, planning_period: str) -> Dict:
        """Get strategic planning data"""
        # Get department performance data
        dept_data = self.get_department_performance_data(department_id, academic_year, 'all')
        teacher_data = self.get_teacher_effectiveness_data(department_id, None, academic_year)
        gap_data = self.get_attainment_gap_data(department_id, None, academic_year)
        
        # Calculate strategic insights
        strategic_insights = {
            'department_id': department_id,
            'academic_year': academic_year,
            'planning_period': planning_period,
            'performance_summary': {
                'overall_performance': dept_data['overall_performance'],
                'total_students': dept_data['total_students'],
                'total_subjects': dept_data['total_subjects'],
                'teacher_effectiveness': teacher_data['summary']['average_effectiveness']
            },
            'strengths': [],
            'weaknesses': [],
            'opportunities': [],
            'threats': [],
            'recommendations': [],
            'action_plan': []
        }
        
        # Analyze strengths and weaknesses
        if dept_data['overall_performance'] >= 70:
            strategic_insights['strengths'].append("Strong overall academic performance")
        else:
            strategic_insights['weaknesses'].append("Below target academic performance")
        
        if teacher_data['summary']['average_effectiveness'] >= 75:
            strategic_insights['strengths'].append("High teacher effectiveness")
        else:
            strategic_insights['weaknesses'].append("Teacher effectiveness needs improvement")
        
        if gap_data['summary']['critical_gaps'] == 0:
            strategic_insights['strengths'].append("No critical attainment gaps")
        else:
            strategic_insights['weaknesses'].append(f"{gap_data['summary']['critical_gaps']} critical attainment gaps")
        
        # Generate recommendations
        if dept_data['overall_performance'] < 70:
            strategic_insights['recommendations'].append("Implement additional student support programs")
        if teacher_data['summary']['average_effectiveness'] < 75:
            strategic_insights['recommendations'].append("Provide teacher training and development programs")
        if gap_data['summary']['critical_gaps'] > 0:
            strategic_insights['recommendations'].append("Address critical CO/PO attainment gaps")
        
        return strategic_insights

    def get_comparative_analysis_data(self, department_id: int, comparison_type: str, time_period: str) -> Dict:
        """Get comparative analysis data"""
        # This would implement comparison logic based on type
        # For now, return a basic structure
        return {
            'department_id': department_id,
            'comparison_type': comparison_type,
            'time_period': time_period,
            'comparison_data': {},
            'insights': []
        }

    def get_accreditation_data(self, department_id: int, academic_year: str, accreditation_standard: str) -> Dict:
        """Get accreditation data"""
        # Get department performance data
        dept_data = self.get_department_performance_data(department_id, academic_year, 'all')
        teacher_data = self.get_teacher_effectiveness_data(department_id, None, academic_year)
        gap_data = self.get_attainment_gap_data(department_id, None, academic_year)
        
        return {
            'department_id': department_id,
            'academic_year': academic_year,
            'accreditation_standard': accreditation_standard,
            'compliance_status': 'Compliant',  # This would be calculated based on standards
            'performance_metrics': {
                'academic_performance': dept_data['overall_performance'],
                'teacher_effectiveness': teacher_data['summary']['average_effectiveness'],
                'attainment_gaps': gap_data['summary']['total_gaps']
            },
            'evidence': {},
            'recommendations': []
        }

    def get_resource_utilization_data(self, department_id: int, resource_type: str, time_period: str) -> Dict:
        """Get resource utilization data"""
        # This would implement resource utilization analysis
        return {
            'department_id': department_id,
            'resource_type': resource_type,
            'time_period': time_period,
            'utilization_data': {},
            'efficiency_metrics': {}
        }

    # Placeholder methods for PDF/Excel/CSV generation
    def generate_class_analytics_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_class_analytics_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_class_analytics_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_co_po_attainment_pdf(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_co_po_attainment_excel(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_co_po_attainment_csv(self, co_data: Dict, po_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_exam_analysis_pdf(self, exam: Any, exam_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_exam_analysis_excel(self, exam: Any, exam_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_exam_analysis_csv(self, exam: Any, exam_data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_comprehensive_analysis_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_comprehensive_analysis_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_comprehensive_analysis_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_performance_trends_report(self, filters: Dict, format_type: str) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_attainment_gaps_report(self, filters: Dict, format_type: str) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_general_analysis_report(self, filters: Dict, format_type: str) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_department_performance_pdf(self, data: Dict, filters: Dict) -> bytes:
        """Generate department performance PDF report"""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Department Performance Report", title_style))
        story.append(Spacer(1, 12))
        
        # Department info
        dept_info = data.get('department', {})
        story.append(Paragraph(f"<b>Department:</b> {dept_info.get('name', 'N/A')}", styles['Normal']))
        story.append(Paragraph(f"<b>Academic Year:</b> {filters.get('academic_year', '2024-25')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Key metrics
        story.append(Paragraph("<b>Key Performance Metrics</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        metrics_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Total Students', str(data.get('total_students', 0)), 'N/A', 'Active'],
            ['Average Performance', f"{data.get('avg_performance', 0):.1f}%", '70%', 'Good' if data.get('avg_performance', 0) >= 70 else 'Needs Improvement'],
            ['CO Attainment', f"{data.get('co_attainment', 0):.1f}%", '70%', 'Good' if data.get('co_attainment', 0) >= 70 else 'Needs Improvement'],
            ['PO Attainment', f"{data.get('po_attainment', 0):.1f}%", '70%', 'Good' if data.get('po_attainment', 0) >= 70 else 'Needs Improvement'],
        ]
        
        table = Table(metrics_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Subject performance
        story.append(Paragraph("<b>Subject-wise Performance</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        subjects = data.get('subject_performance', [])
        if subjects:
            subject_data = [['Subject', 'Code', 'Avg Score', 'Students', 'Status']]
            for subject in subjects:
                status = 'Good' if subject.get('avg_score', 0) >= 70 else 'Needs Improvement'
                subject_data.append([
                    subject.get('name', 'N/A'),
                    subject.get('code', 'N/A'),
                    f"{subject.get('avg_score', 0):.1f}%",
                    str(subject.get('student_count', 0)),
                    status
                ])
            
            subject_table = Table(subject_data)
            subject_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(subject_table)
        else:
            story.append(Paragraph("No subject data available", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_department_performance_excel(self, data: Dict, filters: Dict) -> bytes:
        """Generate department performance Excel report"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Department Performance"
        
        # Title
        ws['A1'] = "Department Performance Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Department info
        dept_info = data.get('department', {})
        ws['A3'] = f"Department: {dept_info.get('name', 'N/A')}"
        ws['A4'] = f"Academic Year: {filters.get('academic_year', '2024-25')}"
        
        # Key metrics
        ws['A6'] = "Key Performance Metrics"
        ws['A6'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Total Students', data.get('total_students', 0), 'N/A', 'Active'],
            ['Average Performance', f"{data.get('avg_performance', 0):.1f}%", '70%', 'Good' if data.get('avg_performance', 0) >= 70 else 'Needs Improvement'],
            ['CO Attainment', f"{data.get('co_attainment', 0):.1f}%", '70%', 'Good' if data.get('co_attainment', 0) >= 70 else 'Needs Improvement'],
            ['PO Attainment', f"{data.get('po_attainment', 0):.1f}%", '70%', 'Good' if data.get('po_attainment', 0) >= 70 else 'Needs Improvement'],
        ]
        
        for row_num, row_data in enumerate(metrics_data, 7):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 7:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Subject performance
        ws['A12'] = "Subject-wise Performance"
        ws['A12'].font = Font(size=14, bold=True)
        
        subjects = data.get('subject_performance', [])
        if subjects:
            subject_headers = ['Subject', 'Code', 'Avg Score', 'Students', 'Status']
            for col_num, header in enumerate(subject_headers, 1):
                cell = ws.cell(row=13, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_num, subject in enumerate(subjects, 14):
                status = 'Good' if subject.get('avg_score', 0) >= 70 else 'Needs Improvement'
                subject_data = [
                    subject.get('name', 'N/A'),
                    subject.get('code', 'N/A'),
                    f"{subject.get('avg_score', 0):.1f}%",
                    subject.get('student_count', 0),
                    status
                ]
                for col_num, value in enumerate(subject_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_department_performance_csv(self, data: Dict, filters: Dict) -> bytes:
        """Generate department performance CSV report"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Department Performance Report"])
        writer.writerow([])
        
        # Department info
        dept_info = data.get('department', {})
        writer.writerow([f"Department: {dept_info.get('name', 'N/A')}"])
        writer.writerow([f"Academic Year: {filters.get('academic_year', '2024-25')}"])
        writer.writerow([])
        
        # Key metrics
        writer.writerow(["Key Performance Metrics"])
        writer.writerow(["Metric", "Value", "Target", "Status"])
        writer.writerow(["Total Students", data.get('total_students', 0), "N/A", "Active"])
        writer.writerow(["Average Performance", f"{data.get('avg_performance', 0):.1f}%", "70%", "Good" if data.get('avg_performance', 0) >= 70 else "Needs Improvement"])
        writer.writerow(["CO Attainment", f"{data.get('co_attainment', 0):.1f}%", "70%", "Good" if data.get('co_attainment', 0) >= 70 else "Needs Improvement"])
        writer.writerow(["PO Attainment", f"{data.get('po_attainment', 0):.1f}%", "70%", "Good" if data.get('po_attainment', 0) >= 70 else "Needs Improvement"])
        writer.writerow([])
        
        # Subject performance
        writer.writerow(["Subject-wise Performance"])
        writer.writerow(["Subject", "Code", "Avg Score", "Students", "Status"])
        
        subjects = data.get('subject_performance', [])
        for subject in subjects:
            status = 'Good' if subject.get('avg_score', 0) >= 70 else 'Needs Improvement'
            writer.writerow([
                subject.get('name', 'N/A'),
                subject.get('code', 'N/A'),
                f"{subject.get('avg_score', 0):.1f}%",
                subject.get('student_count', 0),
                status
            ])
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def generate_teacher_effectiveness_pdf(self, data: Dict, filters: Dict) -> bytes:
        """Generate teacher effectiveness PDF report"""
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
        import io
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        story = []
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=1
        )
        story.append(Paragraph("Teacher Effectiveness Report", title_style))
        story.append(Spacer(1, 12))
        
        # Teacher info
        teacher_info = data.get('teacher', {})
        story.append(Paragraph(f"<b>Teacher:</b> {teacher_info.get('name', 'N/A')}", styles['Normal']))
        story.append(Paragraph(f"<b>Department:</b> {teacher_info.get('department', 'N/A')}", styles['Normal']))
        story.append(Paragraph(f"<b>Academic Year:</b> {filters.get('academic_year', '2024-25')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Effectiveness metrics
        story.append(Paragraph("<b>Effectiveness Metrics</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        metrics_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Overall Rating', f"{data.get('overall_rating', 0):.1f}/5", '4.0', 'Good' if data.get('overall_rating', 0) >= 4.0 else 'Needs Improvement'],
            ['Student Satisfaction', f"{data.get('student_satisfaction', 0):.1f}%", '80%', 'Good' if data.get('student_satisfaction', 0) >= 80 else 'Needs Improvement'],
            ['Class Performance', f"{data.get('class_performance', 0):.1f}%", '70%', 'Good' if data.get('class_performance', 0) >= 70 else 'Needs Improvement'],
            ['CO Attainment', f"{data.get('co_attainment', 0):.1f}%", '70%', 'Good' if data.get('co_attainment', 0) >= 70 else 'Needs Improvement'],
        ]
        
        table = Table(metrics_data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        story.append(table)
        story.append(Spacer(1, 20))
        
        # Subject performance
        story.append(Paragraph("<b>Subject-wise Performance</b>", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        subjects = data.get('subject_performance', [])
        if subjects:
            subject_data = [['Subject', 'Code', 'Avg Score', 'Students', 'Rating']]
            for subject in subjects:
                subject_data.append([
                    subject.get('name', 'N/A'),
                    subject.get('code', 'N/A'),
                    f"{subject.get('avg_score', 0):.1f}%",
                    str(subject.get('student_count', 0)),
                    f"{subject.get('rating', 0):.1f}/5"
                ])
            
            subject_table = Table(subject_data)
            subject_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(subject_table)
        else:
            story.append(Paragraph("No subject data available", styles['Normal']))
        
        doc.build(story)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_teacher_effectiveness_excel(self, data: Dict, filters: Dict) -> bytes:
        """Generate teacher effectiveness Excel report"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Teacher Effectiveness"
        
        # Title
        ws['A1'] = "Teacher Effectiveness Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:E1')
        
        # Teacher info
        teacher_info = data.get('teacher', {})
        ws['A3'] = f"Teacher: {teacher_info.get('name', 'N/A')}"
        ws['A4'] = f"Department: {teacher_info.get('department', 'N/A')}"
        ws['A5'] = f"Academic Year: {filters.get('academic_year', '2024-25')}"
        
        # Effectiveness metrics
        ws['A7'] = "Effectiveness Metrics"
        ws['A7'].font = Font(size=14, bold=True)
        
        metrics_data = [
            ['Metric', 'Value', 'Target', 'Status'],
            ['Overall Rating', f"{data.get('overall_rating', 0):.1f}/5", '4.0', 'Good' if data.get('overall_rating', 0) >= 4.0 else 'Needs Improvement'],
            ['Student Satisfaction', f"{data.get('student_satisfaction', 0):.1f}%", '80%', 'Good' if data.get('student_satisfaction', 0) >= 80 else 'Needs Improvement'],
            ['Class Performance', f"{data.get('class_performance', 0):.1f}%", '70%', 'Good' if data.get('class_performance', 0) >= 70 else 'Needs Improvement'],
            ['CO Attainment', f"{data.get('co_attainment', 0):.1f}%", '70%', 'Good' if data.get('co_attainment', 0) >= 70 else 'Needs Improvement'],
        ]
        
        for row_num, row_data in enumerate(metrics_data, 8):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                if row_num == 8:  # Header row
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # Subject performance
        ws['A13'] = "Subject-wise Performance"
        ws['A13'].font = Font(size=14, bold=True)
        
        subjects = data.get('subject_performance', [])
        if subjects:
            subject_headers = ['Subject', 'Code', 'Avg Score', 'Students', 'Rating']
            for col_num, header in enumerate(subject_headers, 1):
                cell = ws.cell(row=14, column=col_num, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            for row_num, subject in enumerate(subjects, 15):
                subject_data = [
                    subject.get('name', 'N/A'),
                    subject.get('code', 'N/A'),
                    f"{subject.get('avg_score', 0):.1f}%",
                    subject.get('student_count', 0),
                    f"{subject.get('rating', 0):.1f}/5"
                ]
                for col_num, value in enumerate(subject_data, 1):
                    ws.cell(row=row_num, column=col_num, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = None
            for cell in column:
                try:
                    if hasattr(cell, 'column_letter'):
                        if column_letter is None:
                            column_letter = cell.column_letter
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                except:
                    pass
            if column_letter:
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    def generate_teacher_effectiveness_csv(self, data: Dict, filters: Dict) -> bytes:
        """Generate teacher effectiveness CSV report"""
        import csv
        import io
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Teacher Effectiveness Report"])
        writer.writerow([])
        
        # Teacher info
        teacher_info = data.get('teacher', {})
        writer.writerow([f"Teacher: {teacher_info.get('name', 'N/A')}"])
        writer.writerow([f"Department: {teacher_info.get('department', 'N/A')}"])
        writer.writerow([f"Academic Year: {filters.get('academic_year', '2024-25')}"])
        writer.writerow([])
        
        # Effectiveness metrics
        writer.writerow(["Effectiveness Metrics"])
        writer.writerow(["Metric", "Value", "Target", "Status"])
        writer.writerow(["Overall Rating", f"{data.get('overall_rating', 0):.1f}/5", "4.0", "Good" if data.get('overall_rating', 0) >= 4.0 else "Needs Improvement"])
        writer.writerow(["Student Satisfaction", f"{data.get('student_satisfaction', 0):.1f}%", "80%", "Good" if data.get('student_satisfaction', 0) >= 80 else "Needs Improvement"])
        writer.writerow(["Class Performance", f"{data.get('class_performance', 0):.1f}%", "70%", "Good" if data.get('class_performance', 0) >= 70 else "Needs Improvement"])
        writer.writerow(["CO Attainment", f"{data.get('co_attainment', 0):.1f}%", "70%", "Good" if data.get('co_attainment', 0) >= 70 else "Needs Improvement"])
        writer.writerow([])
        
        # Subject performance
        writer.writerow(["Subject-wise Performance"])
        writer.writerow(["Subject", "Code", "Avg Score", "Students", "Rating"])
        
        subjects = data.get('subject_performance', [])
        for subject in subjects:
            writer.writerow([
                subject.get('name', 'N/A'),
                subject.get('code', 'N/A'),
                f"{subject.get('avg_score', 0):.1f}%",
                subject.get('student_count', 0),
                f"{subject.get('rating', 0):.1f}/5"
            ])
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def generate_student_progress_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_student_progress_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_student_progress_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_attainment_gap_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_attainment_gap_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_attainment_gap_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_strategic_planning_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_strategic_planning_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_strategic_planning_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_comparative_analysis_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_comparative_analysis_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_comparative_analysis_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)
    
    def generate_accreditation_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_accreditation_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_resource_utilization_pdf(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_pdf([], None, [], filters)
    
    def generate_resource_utilization_excel(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_excel([], None, [], filters)
    
    def generate_resource_utilization_csv(self, data: Dict, filters: Dict) -> bytes:
        return self.generate_student_performance_csv([], None, [], filters)

    def calculate_gpa(self, percentage: float) -> float:
        """Calculate GPA from percentage"""
        if percentage >= 90:
            return 4.0
        elif percentage >= 80:
            return 3.5
        elif percentage >= 70:
            return 3.0
        elif percentage >= 60:
            return 2.5
        elif percentage >= 50:
            return 2.0
        elif percentage >= 40:
            return 1.5
        else:
            return 0.0

# Additional helper functions for different report types
def get_available_report_types() -> List[Dict[str, Any]]:
    """Get list of available report types"""
    return [
        # Teacher-level reports
        {
            "id": "student_performance",
            "name": "Student Performance Analysis",
            "description": "Detailed analysis of individual student performance",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["subject_id", "class_id", "exam_type"],
            "roles": ["teacher", "hod", "admin"]
        },
        {
            "id": "class_analytics",
            "name": "Class Analytics Report",
            "description": "Comprehensive class-level analytics and statistics",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["subject_id", "class_id", "exam_type"],
            "roles": ["teacher", "hod", "admin"]
        },
        {
            "id": "co_po_attainment",
            "name": "CO/PO Attainment Analysis",
            "description": "Course Outcome and Program Outcome attainment analysis",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["subject_id", "exam_type"],
            "roles": ["teacher", "hod", "admin"]
        },
        {
            "id": "exam_analysis",
            "name": "Exam Analysis Report",
            "description": "Detailed analysis of specific exam performance",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["exam_id", "subject_id"],
            "roles": ["teacher", "hod", "admin"]
        },
        {
            "id": "comprehensive_analysis",
            "name": "Comprehensive Analysis",
            "description": "Complete analysis with all metrics and insights",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["subject_id", "class_id", "exam_type"],
            "roles": ["teacher", "hod", "admin"]
        },
        {
            "id": "custom_analysis",
            "name": "Custom Analysis",
            "description": "Custom analysis based on specific requirements",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["analysis_type", "custom_filters"],
            "roles": ["teacher", "hod", "admin"]
        },
        # HOD-level reports
        {
            "id": "department_performance",
            "name": "Department Performance Report",
            "description": "Comprehensive department-wide performance analysis",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "academic_year", "semester"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "teacher_effectiveness",
            "name": "Teacher Effectiveness Analysis",
            "description": "Analysis of teacher performance and effectiveness",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "teacher_id", "subject_id", "academic_year"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "student_progress_tracking",
            "name": "Student Progress Tracking",
            "description": "Track student progress across subjects and semesters",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "class_id", "student_id", "academic_year"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "attainment_gap_analysis",
            "name": "Attainment Gap Analysis",
            "description": "Identify gaps in CO/PO attainment and improvement areas",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "subject_id", "academic_year"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "strategic_planning",
            "name": "Strategic Planning Report",
            "description": "Strategic insights for department planning and improvement",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "academic_year", "planning_period"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "comparative_analysis",
            "name": "Comparative Analysis Report",
            "description": "Compare performance across classes, subjects, and time periods",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "comparison_type", "time_period"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "accreditation_report",
            "name": "Accreditation Report",
            "description": "Comprehensive report for accreditation and quality assurance",
            "formats": ["pdf", "excel"],
            "filters": ["department_id", "academic_year", "accreditation_standard"],
            "roles": ["hod", "admin"]
        },
        {
            "id": "resource_utilization",
            "name": "Resource Utilization Report",
            "description": "Analysis of resource utilization and efficiency",
            "formats": ["pdf", "excel", "csv"],
            "filters": ["department_id", "resource_type", "time_period"],
            "roles": ["hod", "admin"]
        }
    ]
