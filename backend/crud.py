from sqlalchemy.orm import Session, joinedload
from sqlalchemy import desc, func, and_, or_
from typing import List, Optional, Dict, Any
from models import *
from schemas import *
from auth import get_password_hash
import openpyxl
from io import BytesIO
from datetime import datetime, timedelta

# Utility functions for marks lock-in period
def is_marks_locked(exam_date: datetime) -> bool:
    """Check if marks are locked (more than 7 days after exam date)"""
    if not exam_date:
        return False  # If no exam date, marks are not locked
    
    lock_in_period = timedelta(days=7)
    return datetime.now(exam_date.tzinfo) > exam_date + lock_in_period

def get_lock_in_status(exam_date: datetime) -> Dict[str, Any]:
    """Get detailed lock-in status for an exam"""
    if not exam_date:
        return {
            "is_locked": False,
            "days_remaining": None,
            "lock_date": None,
            "message": "No exam date set"
        }
    
    lock_in_period = timedelta(days=7)
    lock_date = exam_date + lock_in_period
    now = datetime.now(exam_date.tzinfo)
    
    if now > lock_date:
        return {
            "is_locked": True,
            "days_remaining": 0,
            "lock_date": lock_date,
            "message": "Marks are locked - 7-day lock-in period has expired"
        }
    else:
        days_remaining = (lock_date - now).days
        return {
            "is_locked": False,
            "days_remaining": days_remaining,
            "lock_date": lock_date,
            "message": f"Marks can be modified for {days_remaining} more days"
        }

# Department CRUD
def get_all_departments(db: Session):
    return db.query(Department).options(
        joinedload(Department.hod),
        joinedload(Department.classes),
        joinedload(Department.users)
    ).all()

def get_department_by_id(db: Session, department_id: int):
    return db.query(Department).options(
        joinedload(Department.hod),
        joinedload(Department.classes),
        joinedload(Department.users)
    ).filter(Department.id == department_id).first()

def create_department(db: Session, department: DepartmentCreate):
    db_department = Department(**department.dict())
    db.add(db_department)
    db.commit()
    db.refresh(db_department)
    return db_department

def update_department(db: Session, department_id: int, department: DepartmentUpdate):
    db_department = db.query(Department).filter(Department.id == department_id).first()
    if not db_department:
        return None
    
    update_data = department.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_department, field, value)
    
    db.commit()
    db.refresh(db_department)
    return db_department

def delete_department(db: Session, department_id: int):
    db_department = db.query(Department).filter(Department.id == department_id).first()
    if db_department:
        # Check if department has users or classes
        users_count = db.query(User).filter(User.department_id == department_id).count()
        classes_count = db.query(Class).filter(Class.department_id == department_id).count()
        
        if users_count > 0 or classes_count > 0:
            raise ValueError("Cannot delete department with existing users or classes")
        
        db.delete(db_department)
        db.commit()
        return True
    return False

# Class CRUD
def get_all_classes(db: Session):
    return db.query(Class).options(
        joinedload(Class.department),
        joinedload(Class.students),
        joinedload(Class.subjects)
    ).all()

def get_class_by_id(db: Session, class_id: int):
    return db.query(Class).options(
        joinedload(Class.department),
        joinedload(Class.students),
        joinedload(Class.subjects)
    ).filter(Class.id == class_id).first()

def create_class(db: Session, class_data: ClassCreate):
    db_class = Class(**class_data.dict())
    db.add(db_class)
    db.commit()
    db.refresh(db_class)
    return db_class

def update_class(db: Session, class_id: int, class_data: ClassUpdate):
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if not db_class:
        return None
    
    update_data = class_data.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_class, field, value)
    
    db.commit()
    db.refresh(db_class)
    return db_class

def delete_class(db: Session, class_id: int):
    db_class = db.query(Class).filter(Class.id == class_id).first()
    if db_class:
        # Check if class has students or subjects
        students_count = db.query(User).filter(User.class_id == class_id).count()
        subjects_count = db.query(Subject).filter(Subject.class_id == class_id).count()
        
        if students_count > 0 or subjects_count > 0:
            raise ValueError("Cannot delete class with existing students or subjects")
        
        db.delete(db_class)
        db.commit()
        return True
    return False

# Subject CRUD
def get_all_subjects(db: Session):
    return db.query(Subject).options(
        joinedload(Subject.class_obj),
        joinedload(Subject.teacher),
        joinedload(Subject.exams)
    ).all()

def get_subject_by_id(db: Session, subject_id: int):
    return db.query(Subject).options(
        joinedload(Subject.class_obj),
        joinedload(Subject.teacher),
        joinedload(Subject.exams)
    ).filter(Subject.id == subject_id).first()

def create_subject(db: Session, subject: SubjectCreate):
    db_subject = Subject(**subject.dict())
    db.add(db_subject)
    db.commit()
    db.refresh(db_subject)
    return db_subject

def update_subject(db: Session, subject_id: int, subject: SubjectUpdate):
    db_subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not db_subject:
        return None
    
    update_data = subject.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_subject, field, value)
    
    db.commit()
    db.refresh(db_subject)
    return db_subject

def delete_subject(db: Session, subject_id: int):
    db_subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if db_subject:
        # Check if subject has exams
        exams_count = db.query(Exam).filter(Exam.subject_id == subject_id).count()
        
        if exams_count > 0:
            raise ValueError("Cannot delete subject with existing exams")
        
        db.delete(db_subject)
        db.commit()
        return True
    return False

# User CRUD
def get_all_users(db: Session):
    return db.query(User).options(
        joinedload(User.department),
        joinedload(User.student_class)
    ).all()

def get_user_by_id(db: Session, user_id: int):
    return db.query(User).options(
        joinedload(User.department),
        joinedload(User.student_class)
    ).filter(User.id == user_id).first()

def get_user_by_username(db: Session, username: str):
    return db.query(User).filter(User.username == username).first()

def create_user(db: Session, user: UserCreate):
    hashed_password = get_password_hash(user.password)
    user_dict = user.dict()
    user_dict.pop('password')
    
    db_user = User(
        **user_dict,
        hashed_password=hashed_password
    )
    db.add(db_user)
    db.commit()
    db.refresh(db_user)
    return db_user

def update_user(db: Session, user_id: int, user: UserUpdate):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        return None
    
    update_data = user.dict(exclude_unset=True)
    if 'password' in update_data:
        update_data['hashed_password'] = get_password_hash(update_data.pop('password'))
    
    for field, value in update_data.items():
        setattr(db_user, field, value)
    
    db.commit()
    db.refresh(db_user)
    return db_user

def update_user_password(db: Session, user_id: int, new_password: str):
    db_user = db.query(User).filter(User.id == user_id).first()
    if not db_user:
        return False
    
    db_user.hashed_password = get_password_hash(new_password)
    db.commit()
    return True

def delete_user(db: Session, user_id: int):
    db_user = db.query(User).filter(User.id == user_id).first()
    if db_user:
        # Check if user has dependent records
        marks_count = db.query(Mark).filter(Mark.student_id == user_id).count()
        subjects_count = db.query(Subject).filter(Subject.teacher_id == user_id).count()
        
        if marks_count > 0 or subjects_count > 0:
            # Soft delete by deactivating instead of hard delete
            db_user.is_active = False
            db.commit()
        else:
            db.delete(db_user)
            db.commit()
        return True
    return False

# Exam CRUD
def get_all_exams(db: Session):
    return db.query(Exam).options(
        joinedload(Exam.questions),
        joinedload(Exam.subject).joinedload(Subject.class_obj),
        joinedload(Exam.marks)
    ).all()

def get_exam_by_id_db(db: Session, exam_id: int):
    return db.query(Exam).options(
        joinedload(Exam.questions),
        joinedload(Exam.subject).joinedload(Subject.class_obj),
        joinedload(Exam.marks)
    ).filter(Exam.id == exam_id).first()

def create_exam(db: Session, exam: ExamCreate):
    exam_dict = exam.dict()
    questions_data = exam_dict.pop('questions', [])
    
    # Convert exam_date string to datetime if provided
    if exam_dict.get('exam_date'):
        if isinstance(exam_dict['exam_date'], str):
            exam_dict['exam_date'] = datetime.fromisoformat(exam_dict['exam_date'].replace('Z', '+00:00'))
    
    db_exam = Exam(**exam_dict)
    db.add(db_exam)
    db.commit()
    db.refresh(db_exam)
    
    # Create questions
    for question_data in questions_data:
        db_question = Question(exam_id=db_exam.id, **question_data)
        db.add(db_question)
    
    db.commit()
    
    # Reload exam with questions
    return get_exam_by_id_db(db, db_exam.id)

def update_exam(db: Session, exam_id: int, exam: ExamUpdate):
    db_exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not db_exam:
        return None
    
    update_data = exam.dict(exclude_unset=True)
    
    # Handle exam_date conversion
    if 'exam_date' in update_data and update_data['exam_date']:
        if isinstance(update_data['exam_date'], str):
            update_data['exam_date'] = datetime.fromisoformat(update_data['exam_date'].replace('Z', '+00:00'))
    
    # Handle questions separately to avoid SQLAlchemy relationship issues
    questions_data = update_data.pop('questions', None)
    
    for field, value in update_data.items():
        if hasattr(db_exam, field):
            setattr(db_exam, field, value)
    
    # Handle questions update separately
    if questions_data is not None:
        # First, delete all related records in the correct order to avoid foreign key violations
        questions_to_delete = db.query(Question).filter(Question.exam_id == exam_id).all()
        
        for question in questions_to_delete:
            # Delete question_co_weights first
            db.query(QuestionCOWeight).filter(QuestionCOWeight.question_id == question.id).delete()
            # Delete marks
            db.query(Mark).filter(Mark.question_id == question.id).delete()
        
        # Then delete existing questions
        db.query(Question).filter(Question.exam_id == exam_id).delete()
        
        # Add new questions
        for question_data in questions_data:
            question = Question(
                exam_id=exam_id,
                question_number=question_data.get('question_number'),
                question_text=question_data.get('question_text', f"Question {question_data.get('question_number', '')}"),
                max_marks=question_data.get('max_marks'),
                co_mapping=question_data.get('co_mapping', []),
                po_mapping=question_data.get('po_mapping', []),
                section=question_data.get('section'),
                blooms_level=question_data.get('blooms_level'),
                difficulty=question_data.get('difficulty'),
                co_weighting_mode=question_data.get('co_weighting_mode', 'equal_split')
            )
            db.add(question)
    
    db.commit()
    db.refresh(db_exam)
    return db_exam

def delete_exam(db: Session, exam_id: int):
    print(f"CRUD: Attempting to delete exam {exam_id}")
    db_exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if db_exam:
        print(f"CRUD: Found exam {exam_id}, deleting...")
        # First, delete all marks associated with questions for this exam
        questions_to_delete = db.query(Question).filter(Question.exam_id == exam_id).all()
        print(f"CRUD: Found {len(questions_to_delete)} questions to delete")
        for question in questions_to_delete:
            db.query(Mark).filter(Mark.question_id == question.id).delete()
        
        # Then delete questions
        db.query(Question).filter(Question.exam_id == exam_id).delete()
        
        # Finally delete the exam
        db.delete(db_exam)
        db.commit()
        print(f"CRUD: Successfully deleted exam {exam_id}")
        return True
    else:
        print(f"CRUD: Exam {exam_id} not found")
        return False

# Question CRUD
def create_question(db: Session, question: QuestionCreate):
    db_question = Question(**question.dict())
    db.add(db_question)
    db.commit()
    db.refresh(db_question)
    return db_question

def get_questions_by_exam(db: Session, exam_id: int):
    return db.query(Question).filter(Question.exam_id == exam_id).all()

# Mark CRUD
def get_marks_by_exam_id(db: Session, exam_id: int):
    marks = db.query(Mark).options(
        joinedload(Mark.student),
        joinedload(Mark.question),
        joinedload(Mark.exam)
    ).filter(Mark.exam_id == exam_id).all()
    
    # Get exam to check lock-in status
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    lock_status = get_lock_in_status(exam.exam_date) if exam else {"is_locked": False, "message": "Exam not found"}
    
    return {
        "marks": marks,
        "lock_status": lock_status
    }

def get_marks_by_student_id(db: Session, student_id: int):
    return db.query(Mark).options(
        joinedload(Mark.question),
        joinedload(Mark.exam).joinedload(Exam.subject)
    ).filter(Mark.student_id == student_id).all()

def get_student_marks_for_exam(db: Session, exam_id: int, student_id: int):
    return db.query(Mark).options(
        joinedload(Mark.question),
        joinedload(Mark.exam)
    ).filter(
        and_(Mark.exam_id == exam_id, Mark.student_id == student_id)
    ).all()

def bulk_create_marks_db(db: Session, marks: List[MarkCreate]):
    try:
        # Check lock-in status for the first mark's exam (assuming all marks are for the same exam)
        if marks:
            exam = db.query(Exam).filter(Exam.id == marks[0].exam_id).first()
            if exam and is_marks_locked(exam.exam_date):
                raise ValueError("Marks are locked - 7-day lock-in period has expired")
        
        # Group marks by exam/student/question for efficient processing
        for mark_data in marks:
            # Validate mark against question max_marks
            question = db.query(Question).filter(Question.id == mark_data.question_id).first()
            if question and mark_data.marks_obtained > question.max_marks:
                raise ValueError(f"Marks {mark_data.marks_obtained} exceed max marks {question.max_marks} for question {question.question_number}")
            
            # Check if mark already exists
            existing_mark = db.query(Mark).filter(
                and_(
                    Mark.exam_id == mark_data.exam_id,
                    Mark.student_id == mark_data.student_id,
                    Mark.question_id == mark_data.question_id
                )
            ).first()
            
            if existing_mark:
                existing_mark.marks_obtained = mark_data.marks_obtained
            else:
                db_mark = Mark(**mark_data.dict())
                db.add(db_mark)
        
        db.commit()
        
        # Return all marks for the exam
        exam_id = marks[0].exam_id if marks else None
        if exam_id:
            return get_marks_by_exam_id(db, exam_id)
        return []
        
    except Exception as e:
        db.rollback()
        raise e

def get_exam_lock_status(db: Session, exam_id: int):
    """Get lock-in status for a specific exam"""
    exam = db.query(Exam).filter(Exam.id == exam_id).first()
    if not exam:
        return {"error": "Exam not found"}
    
    return get_lock_in_status(exam.exam_date)

def update_mark(db: Session, mark_id: int, mark: MarkUpdate):
    db_mark = db.query(Mark).filter(Mark.id == mark_id).first()
    if not db_mark:
        return None
    
    # Check if marks are locked for this exam
    exam = db.query(Exam).filter(Exam.id == db_mark.exam_id).first()
    if exam and is_marks_locked(exam.exam_date):
        raise ValueError("Marks are locked - 7-day lock-in period has expired")
    
    # Validate mark against question max_marks
    question = db.query(Question).filter(Question.id == db_mark.question_id).first()
    if question and mark.marks_obtained > question.max_marks:
        raise ValueError(f"Marks {mark.marks_obtained} exceed max marks {question.max_marks}")
    
    db_mark.marks_obtained = mark.marks_obtained
    db.commit()
    db.refresh(db_mark)
    return db_mark

# Dashboard stats functions
def get_admin_dashboard_stats(db: Session):
    total_departments = db.query(Department).count()
    total_users = db.query(User).count()
    total_classes = db.query(Class).count()
    total_subjects = db.query(Subject).count()
    active_users = db.query(User).filter(User.is_active == True).count()
    
    return {
        "total_departments": total_departments,
        "total_users": total_users,
        "total_classes": total_classes,
        "total_subjects": total_subjects,
        "active_users": active_users,
        "recent_activity": get_recent_activity(db)
    }

def get_hod_dashboard_stats(db: Session, department_id: int):
    dept_users = db.query(User).filter(User.department_id == department_id)
    total_students = dept_users.filter(User.role == UserRole.student).count()
    total_teachers = dept_users.filter(User.role == UserRole.teacher).count()
    
    dept_classes = db.query(Class).filter(Class.department_id == department_id).count()
    dept_subjects = db.query(Subject).join(Class).filter(Class.department_id == department_id).count()
    
    return {
        "total_students": total_students,
        "total_teachers": total_teachers,
        "total_classes": dept_classes,
        "total_subjects": dept_subjects,
        "department_id": department_id
    }

def get_teacher_dashboard_stats(db: Session, teacher_id: int):
    teacher_subjects = db.query(Subject).filter(Subject.teacher_id == teacher_id).count()
    teacher_exams = db.query(Exam).join(Subject).filter(Subject.teacher_id == teacher_id).count()
    
    # Get students under teacher's subjects
    subject_ids = db.query(Subject.id).filter(Subject.teacher_id == teacher_id).subquery()
    class_ids = db.query(Subject.class_id).filter(Subject.teacher_id == teacher_id).distinct().subquery()
    total_students = db.query(User).filter(
        and_(User.role == UserRole.student, User.class_id.in_(class_ids))
    ).count()
    
    return {
        "subjects_assigned": teacher_subjects,
        "exams_configured": teacher_exams,
        "total_students": total_students,
        "teacher_id": teacher_id
    }

def get_student_dashboard_stats(db: Session, student_id: int):
    student = get_user_by_id(db, student_id)
    if not student:
        return {}
    
    # Get student's marks
    student_marks = get_marks_by_student_id(db, student_id)
    total_exams = len(set(mark.exam_id for mark in student_marks))
    
    # Get class subjects
    class_subjects = db.query(Subject).filter(Subject.class_id == student.class_id).count() if student.class_id else 0
    
    return {
        "total_exams_taken": total_exams,
        "total_subjects": class_subjects,
        "class_id": student.class_id,
        "student_id": student_id
    }

def get_recent_activity(db: Session, limit: int = 10):
    # Get recent users, exams, and marks
    recent_users = db.query(User).order_by(desc(User.created_at)).limit(limit).all()
    recent_exams = db.query(Exam).order_by(desc(Exam.created_at)).limit(limit).all()
    
    activities = []
    
    for user in recent_users[-5:]:  # Last 5
        activities.append({
            "type": "user_created",
            "message": f"New {user.role.value} user '{user.first_name} {user.last_name}' registered",
            "timestamp": user.created_at
        })
    
    for exam in recent_exams[-5:]:  # Last 5
        activities.append({
            "type": "exam_created",
            "message": f"New exam '{exam.name}' created",
            "timestamp": exam.created_at
        })
    
    # Sort by timestamp descending
    activities.sort(key=lambda x: x["timestamp"], reverse=True)
    
    return activities[:limit]

# Excel/File processing functions
def generate_marks_template(db: Session, exam_id: int) -> bytes:
    """Generate Excel template for marks entry"""
    exam = get_exam_by_id_db(db, exam_id)
    if not exam:
        raise ValueError("Exam not found")
    
    # Get students for the exam's class
    subject = get_subject_by_id(db, exam.subject_id)
    if not subject:
        raise ValueError("Subject not found")
    
    students = db.query(User).filter(
        and_(User.role == UserRole.student, User.class_id == subject.class_id)
    ).order_by(User.first_name, User.last_name).all()
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Marks Entry"
    
    # Headers
    headers = ["Student ID", "Student Name", "Email"]
    for question in exam.questions:
        headers.append(f"Q{question.question_number} ({question.max_marks})")
    headers.append("Total")
    
    # Write headers
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Write student data
    for row, student in enumerate(students, 2):
        ws.cell(row=row, column=1, value=student.id)
        ws.cell(row=row, column=2, value=f"{student.first_name} {student.last_name}")
        ws.cell(row=row, column=3, value=student.email)
        
        # Leave question columns empty for data entry
        # Add total formula
        question_cols = len(exam.questions)
        total_col = 4 + question_cols
        if question_cols > 0:
            start_col = openpyxl.utils.get_column_letter(4)
            end_col = openpyxl.utils.get_column_letter(3 + question_cols)
            ws.cell(row=row, column=total_col, value=f"=SUM({start_col}{row}:{end_col}{row})")
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()

async def process_marks_excel_upload(file, exam_id: int, db: Session) -> List[MarkCreate]:
    """Process uploaded Excel file and return marks data"""
    content = await file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active
    
    exam = get_exam_by_id_db(db, exam_id)
    if not exam:
        raise ValueError("Exam not found")
    
    marks_data = []
    
    # Read headers to identify question columns
    headers = [cell.value for cell in ws[1]]
    question_columns = {}
    
    for i, header in enumerate(headers):
        if header and header.startswith('Q'):
            # Extract question number from header like "Q1a (10)"
            question_num = header.split('(')[0].replace('Q', '').strip()
            question = db.query(Question).filter(
                and_(Question.exam_id == exam_id, Question.question_number == question_num)
            ).first()
            if question:
                question_columns[i] = question.id
    
    # Process each row (skip header)
    for row_num in range(2, ws.max_row + 1):
        student_id = ws.cell(row=row_num, column=1).value
        if not student_id:
            continue
            
        for col_idx, question_id in question_columns.items():
            marks_value = ws.cell(row=row_num, column=col_idx + 1).value
            if marks_value is not None:
                try:
                    marks_obtained = float(marks_value)
                    marks_data.append(MarkCreate(
                        exam_id=exam_id,
                        student_id=int(student_id),
                        question_id=question_id,
                        marks_obtained=marks_obtained
                    ))
                except (ValueError, TypeError):
                    continue  # Skip invalid marks
    
    return marks_data

# CO/PO/PSO Framework CRUD Operations

# CO Definition CRUD
def get_co_definitions_by_subject(db: Session, subject_id: int) -> List[CODefinition]:
    """Get all CO definitions for a subject"""
    return db.query(CODefinition).filter(CODefinition.subject_id == subject_id).all()

def get_co_definition_by_id(db: Session, co_id: int) -> Optional[CODefinition]:
    """Get CO definition by ID"""
    return db.query(CODefinition).filter(CODefinition.id == co_id).first()

def create_co_definition(db: Session, co_definition: CODefinitionCreate) -> CODefinition:
    """Create a new CO definition"""
    db_co = CODefinition(**co_definition.dict())
    db.add(db_co)
    db.commit()
    db.refresh(db_co)
    return db_co

def update_co_definition(db: Session, co_id: int, co_definition: CODefinitionUpdate) -> Optional[CODefinition]:
    """Update CO definition"""
    db_co = db.query(CODefinition).filter(CODefinition.id == co_id).first()
    if not db_co:
        return None
    
    update_data = co_definition.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_co, field, value)
    
    db.commit()
    db.refresh(db_co)
    return db_co

def delete_co_definition(db: Session, co_id: int) -> bool:
    """Delete CO definition"""
    db_co = db.query(CODefinition).filter(CODefinition.id == co_id).first()
    if db_co:
        db.delete(db_co)
        db.commit()
        return True
    return False

# PO Definition CRUD
def get_po_definitions_by_department(db: Session, department_id: int) -> List[PODefinition]:
    """Get all PO definitions for a department"""
    return db.query(PODefinition).filter(PODefinition.department_id == department_id).all()

def get_po_definition_by_id(db: Session, po_id: int) -> Optional[PODefinition]:
    """Get PO definition by ID"""
    return db.query(PODefinition).filter(PODefinition.id == po_id).first()

def create_po_definition(db: Session, po_definition: PODefinitionCreate) -> PODefinition:
    """Create a new PO definition"""
    db_po = PODefinition(**po_definition.dict())
    db.add(db_po)
    db.commit()
    db.refresh(db_po)
    return db_po

def update_po_definition(db: Session, po_id: int, po_definition: PODefinitionUpdate) -> Optional[PODefinition]:
    """Update PO definition"""
    db_po = db.query(PODefinition).filter(PODefinition.id == po_id).first()
    if not db_po:
        return None
    
    update_data = po_definition.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_po, field, value)
    
    db.commit()
    db.refresh(db_po)
    return db_po

def delete_po_definition(db: Session, po_id: int) -> bool:
    """Delete PO definition"""
    db_po = db.query(PODefinition).filter(PODefinition.id == po_id).first()
    if db_po:
        db.delete(db_po)
        db.commit()
        return True
    return False

# CO Target CRUD
def get_co_targets_by_subject(db: Session, subject_id: int) -> List[COTarget]:
    """Get all CO targets for a subject"""
    return db.query(COTarget).filter(COTarget.subject_id == subject_id).all()

def get_co_target_by_id(db: Session, target_id: int) -> Optional[COTarget]:
    """Get CO target by ID"""
    return db.query(COTarget).filter(COTarget.id == target_id).first()

def create_co_target(db: Session, co_target: COTargetCreate) -> COTarget:
    """Create a new CO target"""
    db_target = COTarget(**co_target.dict())
    db.add(db_target)
    db.commit()
    db.refresh(db_target)
    return db_target

def update_co_target(db: Session, target_id: int, co_target: COTargetUpdate) -> Optional[COTarget]:
    """Update CO target"""
    db_target = db.query(COTarget).filter(COTarget.id == target_id).first()
    if not db_target:
        return None
    
    update_data = co_target.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_target, field, value)
    
    db.commit()
    db.refresh(db_target)
    return db_target

def delete_co_target(db: Session, target_id: int) -> bool:
    """Delete CO target"""
    db_target = db.query(COTarget).filter(COTarget.id == target_id).first()
    if db_target:
        db.delete(db_target)
        db.commit()
        return True
    return False

def bulk_update_co_targets(db: Session, subject_id: int, co_targets: List[COTargetCreate]) -> List[COTarget]:
    """Bulk update CO targets for a subject"""
    # Delete existing targets
    db.query(COTarget).filter(COTarget.subject_id == subject_id).delete()
    
    # Create new targets
    new_targets = []
    for co_target in co_targets:
        # Verify CO exists and belongs to subject
        co_definition = db.query(CODefinition).filter(
            CODefinition.id == co_target.co_id,
            CODefinition.subject_id == subject_id
        ).first()
        
        if not co_definition:
            raise ValueError(f"CO definition with ID '{co_target.co_id}' not found for subject {subject_id}")
        
        # Create COTarget with co_id
        db_target = COTarget(
            subject_id=subject_id,
            co_id=co_target.co_id,
            target_pct=co_target.target_pct,
            l1_threshold=co_target.l1_threshold,
            l2_threshold=co_target.l2_threshold,
            l3_threshold=co_target.l3_threshold
        )
        db.add(db_target)
        new_targets.append(db_target)
    
    db.commit()
    for target in new_targets:
        db.refresh(target)
    
    return new_targets

# Assessment Weight CRUD
def get_assessment_weights_by_subject(db: Session, subject_id: int) -> List[AssessmentWeight]:
    """Get all assessment weights for a subject"""
    return db.query(AssessmentWeight).filter(AssessmentWeight.subject_id == subject_id).all()

def get_assessment_weight_by_id(db: Session, weight_id: int) -> Optional[AssessmentWeight]:
    """Get assessment weight by ID"""
    return db.query(AssessmentWeight).filter(AssessmentWeight.id == weight_id).first()

def create_assessment_weight(db: Session, assessment_weight: AssessmentWeightCreate) -> AssessmentWeight:
    """Create a new assessment weight"""
    db_weight = AssessmentWeight(**assessment_weight.dict())
    db.add(db_weight)
    db.commit()
    db.refresh(db_weight)
    return db_weight

def update_assessment_weight(db: Session, weight_id: int, assessment_weight: AssessmentWeightUpdate) -> Optional[AssessmentWeight]:
    """Update assessment weight"""
    db_weight = db.query(AssessmentWeight).filter(AssessmentWeight.id == weight_id).first()
    if not db_weight:
        return None
    
    update_data = assessment_weight.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_weight, field, value)
    
    db.commit()
    db.refresh(db_weight)
    return db_weight

def delete_assessment_weight(db: Session, weight_id: int) -> bool:
    """Delete assessment weight"""
    db_weight = db.query(AssessmentWeight).filter(AssessmentWeight.id == weight_id).first()
    if db_weight:
        db.delete(db_weight)
        db.commit()
        return True
    return False

def bulk_update_assessment_weights(db: Session, subject_id: int, assessment_weights: List[AssessmentWeightCreate]) -> List[AssessmentWeight]:
    """Bulk update assessment weights for a subject"""
    # Delete existing weights
    db.query(AssessmentWeight).filter(AssessmentWeight.subject_id == subject_id).delete()
    
    # Create new weights
    new_weights = []
    for weight in assessment_weights:
        weight.subject_id = subject_id
        db_weight = AssessmentWeight(**weight.dict())
        db.add(db_weight)
        new_weights.append(db_weight)
    
    db.commit()
    for weight in new_weights:
        db.refresh(weight)
    
    return new_weights

# CO-PO Matrix CRUD
def get_co_po_matrix_by_subject(db: Session, subject_id: int) -> List[COPOMatrix]:
    """Get all CO-PO matrix entries for a subject"""
    return db.query(COPOMatrix).filter(COPOMatrix.subject_id == subject_id).all()

def get_co_po_matrix_by_id(db: Session, matrix_id: int) -> Optional[COPOMatrix]:
    """Get CO-PO matrix entry by ID"""
    return db.query(COPOMatrix).filter(COPOMatrix.id == matrix_id).first()

def create_co_po_matrix(db: Session, co_po_matrix: COPOMatrixCreate) -> COPOMatrix:
    """Create a new CO-PO matrix entry"""
    db_matrix = COPOMatrix(**co_po_matrix.dict())
    db.add(db_matrix)
    db.commit()
    db.refresh(db_matrix)
    return db_matrix

def update_co_po_matrix(db: Session, matrix_id: int, co_po_matrix: COPOMatrixUpdate) -> Optional[COPOMatrix]:
    """Update CO-PO matrix entry"""
    db_matrix = db.query(COPOMatrix).filter(COPOMatrix.id == matrix_id).first()
    if not db_matrix:
        return None
    
    update_data = co_po_matrix.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_matrix, field, value)
    
    db.commit()
    db.refresh(db_matrix)
    return db_matrix

def delete_co_po_matrix(db: Session, matrix_id: int) -> bool:
    """Delete CO-PO matrix entry"""
    db_matrix = db.query(COPOMatrix).filter(COPOMatrix.id == matrix_id).first()
    if db_matrix:
        db.delete(db_matrix)
        db.commit()
        return True
    return False

def bulk_update_co_po_matrix(db: Session, subject_id: int, co_po_matrix: List[COPOMatrixCreate]) -> List[COPOMatrix]:
    """Bulk update CO-PO matrix for a subject"""
    # Delete existing matrix entries
    db.query(COPOMatrix).filter(COPOMatrix.subject_id == subject_id).delete()
    
    # Create new matrix entries
    new_entries = []
    for entry in co_po_matrix:
        # Verify CO exists and belongs to subject
        co_definition = db.query(CODefinition).filter(
            CODefinition.id == entry.co_id,
            CODefinition.subject_id == subject_id
        ).first()
        
        if not co_definition:
            raise ValueError(f"CO definition with ID '{entry.co_id}' not found for subject {subject_id}")
        
        # Verify PO exists
        po_definition = db.query(PODefinition).filter(
            PODefinition.id == entry.po_id
        ).first()
        
        if not po_definition:
            raise ValueError(f"PO definition with ID '{entry.po_id}' not found")
        
        # Create COPOMatrix with co_id and po_id
        db_entry = COPOMatrix(
            subject_id=subject_id,
            co_id=entry.co_id,
            po_id=entry.po_id,
            strength=entry.strength
        )
        db.add(db_entry)
        new_entries.append(db_entry)
    
    db.commit()
    for entry in new_entries:
        db.refresh(entry)
    
    return new_entries

# Question CO Weight CRUD
def get_question_co_weights(db: Session, question_id: int) -> List[QuestionCOWeight]:
    """Get all CO weights for a question"""
    return db.query(QuestionCOWeight).filter(QuestionCOWeight.question_id == question_id).all()

def get_question_co_weight_by_id(db: Session, weight_id: int) -> Optional[QuestionCOWeight]:
    """Get question CO weight by ID"""
    return db.query(QuestionCOWeight).filter(QuestionCOWeight.id == weight_id).first()

def create_question_co_weight(db: Session, question_co_weight: QuestionCOWeightCreate) -> QuestionCOWeight:
    """Create a new question CO weight"""
    db_weight = QuestionCOWeight(**question_co_weight.dict())
    db.add(db_weight)
    db.commit()
    db.refresh(db_weight)
    return db_weight

def update_question_co_weight(db: Session, weight_id: int, question_co_weight: QuestionCOWeightUpdate) -> Optional[QuestionCOWeight]:
    """Update question CO weight"""
    db_weight = db.query(QuestionCOWeight).filter(QuestionCOWeight.id == weight_id).first()
    if not db_weight:
        return None
    
    update_data = question_co_weight.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_weight, field, value)
    
    db.commit()
    db.refresh(db_weight)
    return db_weight

def delete_question_co_weight(db: Session, weight_id: int) -> bool:
    """Delete question CO weight"""
    db_weight = db.query(QuestionCOWeight).filter(QuestionCOWeight.id == weight_id).first()
    if db_weight:
        db.delete(db_weight)
        db.commit()
        return True
    return False

def bulk_update_question_co_weights(db: Session, question_id: int, co_weights: List[QuestionCOWeightCreate]) -> List[QuestionCOWeight]:
    """Bulk update CO weights for a question"""
    # Delete existing weights
    db.query(QuestionCOWeight).filter(QuestionCOWeight.question_id == question_id).delete()
    
    # Create new weights
    new_weights = []
    for weight in co_weights:
        # Verify CO exists
        co_definition = db.query(CODefinition).filter(CODefinition.id == weight.co_id).first()
        if not co_definition:
            raise ValueError(f"CO definition with ID '{weight.co_id}' not found")
        
        weight.question_id = question_id
        db_weight = QuestionCOWeight(**weight.dict())
        db.add(db_weight)
        new_weights.append(db_weight)
    
    db.commit()
    for weight in new_weights:
        db.refresh(weight)
    
    return new_weights

# Indirect Attainment CRUD
def get_indirect_attainment_by_subject(db: Session, subject_id: int) -> List[IndirectAttainment]:
    """Get all indirect attainment data for a subject"""
    return db.query(IndirectAttainment).filter(IndirectAttainment.subject_id == subject_id).all()

def get_indirect_attainment_by_id(db: Session, attainment_id: int) -> Optional[IndirectAttainment]:
    """Get indirect attainment by ID"""
    return db.query(IndirectAttainment).filter(IndirectAttainment.id == attainment_id).first()

def create_indirect_attainment(db: Session, indirect_attainment: IndirectAttainmentCreate) -> IndirectAttainment:
    """Create a new indirect attainment entry"""
    db_attainment = IndirectAttainment(**indirect_attainment.dict())
    db.add(db_attainment)
    db.commit()
    db.refresh(db_attainment)
    return db_attainment

def update_indirect_attainment(db: Session, attainment_id: int, indirect_attainment: IndirectAttainmentUpdate) -> Optional[IndirectAttainment]:
    """Update indirect attainment"""
    db_attainment = db.query(IndirectAttainment).filter(IndirectAttainment.id == attainment_id).first()
    if not db_attainment:
        return None
    
    update_data = indirect_attainment.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_attainment, field, value)
    
    db.commit()
    db.refresh(db_attainment)
    return db_attainment

def delete_indirect_attainment(db: Session, attainment_id: int) -> bool:
    """Delete indirect attainment"""
    db_attainment = db.query(IndirectAttainment).filter(IndirectAttainment.id == attainment_id).first()
    if db_attainment:
        db.delete(db_attainment)
        db.commit()
        return True
    return False

# Attainment Audit CRUD
def create_attainment_audit(db: Session, subject_id: int, change: str, user_id: int) -> AttainmentAudit:
    """Create an audit entry for attainment changes"""
    db_audit = AttainmentAudit(
        subject_id=subject_id,
        change=change,
        user_id=user_id
    )
    db.add(db_audit)
    db.commit()
    db.refresh(db_audit)
    return db_audit

def get_attainment_audit_by_subject(db: Session, subject_id: int) -> List[AttainmentAudit]:
    """Get audit trail for a subject"""
    return db.query(AttainmentAudit).filter(AttainmentAudit.subject_id == subject_id).order_by(AttainmentAudit.timestamp.desc()).all()

# CO-PO Analysis Functions
def get_co_po_matrix_by_subject(db: Session, subject_id: int) -> List[Dict]:
    """Get CO-PO mapping matrix for a subject"""
    co_po_entries = db.query(COPOMatrix).options(
        joinedload(COPOMatrix.co_definition),
        joinedload(COPOMatrix.po_definition)
    ).filter(COPOMatrix.subject_id == subject_id).all()
    
    # Group by CO
    co_po_map = {}
    for entry in co_po_entries:
        co_code = entry.co_definition.code
        if co_code not in co_po_map:
            co_po_map[co_code] = {
                'co_code': co_code,
                'co_title': entry.co_definition.title,
                'mapped_pos': []
            }
        co_po_map[co_code]['mapped_pos'].append({
            'po_code': entry.po_definition.code,
            'po_title': entry.po_definition.title,
            'strength': entry.strength
        })
    
    return list(co_po_map.values())

def get_co_attainment_analysis_db(db: Session, subject_id: int, exam_type: str = "all") -> Dict:
    """Get CO attainment analysis for a subject"""
    # Get all COs for the subject
    cos = db.query(CODefinition).filter(CODefinition.subject_id == subject_id).all()
    
    # Get exams based on type
    exam_query = db.query(Exam).filter(Exam.subject_id == subject_id)
    if exam_type != "all":
        exam_query = exam_query.filter(Exam.exam_type == exam_type)
    exams = exam_query.all()
    
    co_analysis = {}
    for co in cos:
        co_marks = []
        total_marks = 0
        target_marks = 0
        
        # Get CO target
        co_target = db.query(COTarget).filter(
            COTarget.subject_id == subject_id,
            COTarget.co_id == co.id
        ).first()
        
        if co_target:
            target_marks = co_target.target_pct
        
        # Calculate attainment for each exam
        for exam in exams:
            # Get all marks for the exam and filter by CO mapping in Python
            exam_marks = db.query(Mark).join(Question).filter(
                Mark.exam_id == exam.id
            ).all()
            
            # Filter marks where the question's co_mapping contains the CO code
            exam_marks = [mark for mark in exam_marks 
                         if mark.question.co_mapping and co.code in mark.question.co_mapping]
            
            if exam_marks:
                total_obtained = sum(mark.marks_obtained for mark in exam_marks)
                total_possible = sum(question.max_marks for question in exam.questions 
                                  if co.code in (question.co_mapping or []))
                
                if total_possible > 0:
                    attainment = (total_obtained / total_possible) * 100
                    co_marks.append({
                        'exam_id': exam.id,
                        'exam_name': exam.name,
                        'attainment': attainment,
                        'obtained': total_obtained,
                        'possible': total_possible
                    })
        
        # Calculate overall attainment
        overall_attainment = 0
        if co_marks:
            overall_attainment = sum(m['attainment'] for m in co_marks) / len(co_marks)
        
        co_analysis[co.code] = {
            'co_code': co.code,
            'co_title': co.title,
            'target': target_marks,
            'attainment': overall_attainment,
            'status': 'Achieved' if overall_attainment >= target_marks else 'Not Achieved',
            'exam_details': co_marks
        }
    
    return co_analysis

def get_po_attainment_analysis_db(db: Session, subject_id: int, exam_type: str = "all") -> Dict:
    """Get PO attainment analysis for a subject"""
    # Get CO-PO mappings for the subject
    co_po_matrix = db.query(COPOMatrix).filter(COPOMatrix.subject_id == subject_id).all()
    
    # Get CO attainments
    co_attainments = get_co_attainment_analysis_db(db, subject_id, exam_type)
    
    # Group POs by their mapped COs
    po_analysis = {}
    for entry in co_po_matrix:
        po_code = entry.po_definition.code
        co_code = entry.co_definition.code
        strength = entry.strength
        
        if po_code not in po_analysis:
            po_analysis[po_code] = {
                'po_code': po_code,
                'mapped_cos': [],
                'weighted_attainment': 0,
                'total_weight': 0
            }
        
        if co_code in co_attainments:
            co_attainment = co_attainments[co_code]['attainment']
            po_analysis[po_code]['mapped_cos'].append({
                'co_code': co_code,
                'strength': strength,
                'attainment': co_attainment
            })
            
            # Calculate weighted attainment
            po_analysis[po_code]['weighted_attainment'] += co_attainment * strength
            po_analysis[po_code]['total_weight'] += strength
    
    # Calculate final PO attainments
    for po_code, data in po_analysis.items():
        if data['total_weight'] > 0:
            data['attainment'] = data['weighted_attainment'] / data['total_weight']
        else:
            data['attainment'] = 0
        
        data['status'] = 'Achieved' if data['attainment'] >= 60 else 'Not Achieved'
    
    return po_analysis

def get_student_performance_analysis_db(db: Session, subject_id: int, student_id: int = None, exam_type: str = "all") -> Dict:
    """Get detailed student performance analysis"""
    # Get subject and class
    subject = db.query(Subject).filter(Subject.id == subject_id).first()
    if not subject:
        return {}
    
    # Get students (all or specific)
    if student_id:
        students = db.query(User).filter(User.id == student_id, User.role == 'student').all()
    else:
        students = db.query(User).filter(
            User.class_id == subject.class_id, 
            User.role == 'student'
        ).all()
    
    # Get exams
    exam_query = db.query(Exam).filter(Exam.subject_id == subject_id)
    if exam_type != "all":
        exam_query = exam_query.filter(Exam.exam_type == exam_type)
    exams = exam_query.all()
    
    student_analysis = {}
    for student in students:
        student_marks = []
        total_obtained = 0
        total_possible = 0
        
        for exam in exams:
            exam_marks = db.query(Mark).join(Question).filter(
                Mark.exam_id == exam.id,
                Mark.student_id == student.id
            ).all()
            
            if exam_marks:
                obtained = sum(mark.marks_obtained for mark in exam_marks)
                possible = sum(question.max_marks for question in exam.questions)
                
                student_marks.append({
                    'exam_id': exam.id,
                    'exam_name': exam.name,
                    'obtained': obtained,
                    'possible': possible,
                    'percentage': (obtained / possible * 100) if possible > 0 else 0
                })
                
                total_obtained += obtained
                total_possible += possible
        
        overall_percentage = (total_obtained / total_possible * 100) if total_possible > 0 else 0
        
        student_analysis[student.id] = {
            'student_id': student.id,
            'student_name': f"{student.first_name} {student.last_name}",
            'overall_percentage': overall_percentage,
            'total_obtained': total_obtained,
            'total_possible': total_possible,
            'exam_details': student_marks,
            'grade': get_grade_from_percentage(overall_percentage)
        }
    
    return student_analysis

def get_class_performance_analysis_db(db: Session, subject_id: int, exam_type: str = "all") -> Dict:
    """Get class performance analysis with comparative charts"""
    # Get student performance
    student_analysis = get_student_performance_analysis_db(db, subject_id, None, exam_type)
    
    if not student_analysis:
        return {}
    
    # Calculate class statistics
    percentages = [data['overall_percentage'] for data in student_analysis.values()]
    
    class_stats = {
        'total_students': len(student_analysis),
        'average_percentage': sum(percentages) / len(percentages) if percentages else 0,
        'highest_percentage': max(percentages) if percentages else 0,
        'lowest_percentage': min(percentages) if percentages else 0,
        'pass_rate': len([p for p in percentages if p >= 50]) / len(percentages) * 100 if percentages else 0,
        'grade_distribution': get_grade_distribution(percentages),
        'performance_trends': get_performance_trends(student_analysis),
        'student_rankings': sorted(student_analysis.values(), key=lambda x: x['overall_percentage'], reverse=True)
    }
    
    return class_stats

def get_grade_from_percentage(percentage: float) -> str:
    """Convert percentage to grade"""
    if percentage >= 90:
        return 'A+'
    elif percentage >= 80:
        return 'A'
    elif percentage >= 70:
        return 'B+'
    elif percentage >= 60:
        return 'B'
    elif percentage >= 50:
        return 'C+'
    elif percentage >= 40:
        return 'C'
    else:
        return 'F'

def get_grade_distribution(percentages: List[float]) -> Dict[str, int]:
    """Get grade distribution from percentages"""
    distribution = {'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'F': 0}
    
    for percentage in percentages:
        grade = get_grade_from_percentage(percentage)
        distribution[grade] += 1
    
    return distribution

def get_performance_trends(student_analysis: Dict) -> Dict:
    """Get performance trends across exams"""
    # This would analyze performance trends across different exams
    # For now, return basic structure
    return {
        'improving_students': [],
        'declining_students': [],
        'consistent_students': []
    }

# Student Goals CRUD
def get_student_goals(db: Session, student_id: int) -> List[StudentGoal]:
    """Get all goals for a student"""
    return db.query(StudentGoal).filter(StudentGoal.student_id == student_id).all()

def get_student_goal_by_id(db: Session, goal_id: int, student_id: int) -> Optional[StudentGoal]:
    """Get a specific goal for a student"""
    return db.query(StudentGoal).filter(
        StudentGoal.id == goal_id,
        StudentGoal.student_id == student_id
    ).first()

def create_student_goal(db: Session, goal: StudentGoalCreate) -> StudentGoal:
    """Create a new goal for a student"""
    db_goal = StudentGoal(**goal.dict())
    db.add(db_goal)
    db.commit()
    db.refresh(db_goal)
    return db_goal

def update_student_goal(db: Session, goal_id: int, student_id: int, goal_update: StudentGoalUpdate) -> Optional[StudentGoal]:
    """Update a student's goal"""
    db_goal = get_student_goal_by_id(db, goal_id, student_id)
    if not db_goal:
        return None
    
    update_data = goal_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_goal, field, value)
    
    db.commit()
    db.refresh(db_goal)
    return db_goal

def delete_student_goal(db: Session, goal_id: int, student_id: int) -> bool:
    """Delete a student's goal"""
    db_goal = get_student_goal_by_id(db, goal_id, student_id)
    if not db_goal:
        return False
    
    db.delete(db_goal)
    db.commit()
    return True

# Student Milestones CRUD
def get_student_milestones(db: Session, student_id: int) -> List[StudentMilestone]:
    """Get all milestones for a student"""
    return db.query(StudentMilestone).filter(StudentMilestone.student_id == student_id).all()

def get_student_milestone_by_id(db: Session, milestone_id: int, student_id: int) -> Optional[StudentMilestone]:
    """Get a specific milestone for a student"""
    return db.query(StudentMilestone).filter(
        StudentMilestone.id == milestone_id,
        StudentMilestone.student_id == student_id
    ).first()

def create_student_milestone(db: Session, milestone: StudentMilestoneCreate) -> StudentMilestone:
    """Create a new milestone for a student"""
    db_milestone = StudentMilestone(**milestone.dict())
    db.add(db_milestone)
    db.commit()
    db.refresh(db_milestone)
    return db_milestone

def update_student_milestone(db: Session, milestone_id: int, student_id: int, milestone_update: StudentMilestoneUpdate) -> Optional[StudentMilestone]:
    """Update a student's milestone"""
    db_milestone = get_student_milestone_by_id(db, milestone_id, student_id)
    if not db_milestone:
        return None
    
    update_data = milestone_update.dict(exclude_unset=True)
    for field, value in update_data.items():
        setattr(db_milestone, field, value)
    
    db.commit()
    db.refresh(db_milestone)
    return db_milestone

def delete_student_milestone(db: Session, milestone_id: int, student_id: int) -> bool:
    """Delete a student's milestone"""
    db_milestone = get_student_milestone_by_id(db, milestone_id, student_id)
    if not db_milestone:
        return False
    
    db.delete(db_milestone)
    db.commit()
    return True

def get_student_progress_data(db: Session, student_id: int) -> Dict[str, Any]:
    """Get comprehensive progress data for a student including goals, milestones, and analytics"""
    from analytics import get_student_analytics
    
    goals = get_student_goals(db, student_id)
    milestones = get_student_milestones(db, student_id)
    analytics = get_student_analytics(db, student_id)
    
    return {
        'goals': goals,
        'milestones': milestones,
        'analytics': analytics
    }